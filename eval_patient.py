"""
eval_patient.py — финальный сквозной тест на синтетическом наборе: реальный
пайплайн (pre-evaluation gate -> LLM-судьи) на ВСЕХ суммаризациях одного
пациента, с ранжированием и поиском ЛУЧШЕЙ суммаризации.

Это НЕ регрессия synthetic.py (там эталон vs искажение). Здесь — прод-сценарий
«дан исходник ЭМК и набор суммаризаций-кандидатов, выбери лучшую»: каждая
суммаризация (строки 1..38: эталон + 37 искажённых) сверяется с ИСХОДНИКОМ
(строка 0) тем же конвейером, что и в проде.

ВАЖНО (см. диагностику): gate со scope=None отклоняет ЛЮБУЮ сжатую
суммаризацию по entity_recall (она закономерно не покрывает >=50% сущностей
полного исходника; порог калиброван под scope-aware прод-корпус). Поэтому gate
здесь — информативный ПЕРВЫЙ ЭТАП (его вердикт и причины фиксируются в отчёте),
но дискриминацию «лучше/хуже» даёт ЭТАП СУДЕЙ, который мы прогоняем на всех
кандидатах. Это честно отражено в отчёте.

Запуск:
    python eval_patient.py [--patient Ж1] [--limit-rows N] [--out PREFIX]
"""

from __future__ import annotations

import argparse
import glob
import json
import logging
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd

import config
import objective_layer
import gate
import judge
from llm_client import JudgePanel, LLMError
from synthetic import ROW_DISTORTIONS, SHEET_NAME, ROW_LABEL_COL  # легенда искажений

log = logging.getLogger("eval_patient")

# Порядок предпочтения итоговых категорий (для ранжирования)
CATEGORY_RANK = {
    "Готово к клиническому применению": 3,
    "Требует редактирования": 2,
    "Неприемлемо": 1,
    "ошибка": 0,
}

# Подпапка чекпоинтов — переопределяется в main() при --no-preprocessing, чтобы
# прогон «только LLM-судьи» НЕ смешивался (и не переиспользовал как кэш) со
# старыми grounded-результатами того же пациента.
CKPT_SUBDIR = "eval_patient"


def _ckpt_dir() -> Path:
    return Path("checkpoints") / CKPT_SUBDIR


@dataclass
class SummaryEval:
    """Результат прогона одной суммаризации-кандидата через полный пайплайн."""
    row: int
    distortion_type: str
    severity: str            # critical | benign | gold | "-"
    chars: int
    gate_status: str
    gate_reasons: list[str]
    category: str
    e1_triggered: bool
    block_pass: dict[str, str]      # {"A": "3/3", ...} — сколько подкритериев pass
    objective_findings: list[str]
    verdict: str

    @property
    def n_objective(self) -> int:
        return len(self.objective_findings)

    @property
    def real_error_findings(self) -> list[str]:
        """«Жёсткие» находки, означающие РЕАЛЬНУЮ ошибку (искажённое число/единица,
        инверсия полярности, ложная причинность/интерпретация) — в отличие от
        «пропущенных сущностей», которые для сводки vs полный источник суть
        закономерная КОМПРЕССИЯ, а не ошибка."""
        return [f for f in self.objective_findings if "не найдено" not in f]

    @property
    def n_real_errors(self) -> int:
        return len(self.real_error_findings)

    @property
    def n_compression(self) -> int:
        return sum(1 for f in self.objective_findings if "не найдено" in f)

    @property
    def total_pass(self) -> int:
        total = 0
        for v in self.block_pass.values():
            num = v.split("/")[0]
            total += int(num) if num.isdigit() else 0
        return total

    def rank_key(self) -> tuple:
        """Ключ сортировки: лучше = выше. Первично — категория судей; затем нет E1;
        затем меньше РЕАЛЬНЫХ ошибок объективного слоя (искажённые числа/полярность/
        причинность/интерпретация + галлюцинации/подмены сущностей). Компрессия
        (пропущенные сущности) в ключ НЕ входит — она ~постоянна и зашумлена
        недетерминизмом извлечения, раньше из-за неё эталон (с наибольшим числом
        законных пропусков) падал в хвост. Финальный тай-брейк — пройденные
        подкритерии судей."""
        return (
            CATEGORY_RANK.get(self.category, 0),
            0 if self.e1_triggered else 1,
            -self.n_real_errors,
            self.total_pass,
        )


def _load_checkpoint(patient: str, row: int) -> Optional["SummaryEval"]:
    """Если строка уже посчитана в прошлом (прерванном) прогоне — восстанавливаем
    результат из чекпоинта и не гоняем судей повторно."""
    ckpt = _ckpt_dir() / f"{patient}__row{row}.json"
    if not ckpt.exists():
        return None
    try:
        d = json.loads(ckpt.read_text(encoding="utf-8"))
        return SummaryEval(
            row=d["row"], distortion_type=d["dtype"], severity=d["severity"],
            chars=d["chars"], gate_status=d["gate_status"], gate_reasons=d["gate_reasons"],
            category=d["category"], e1_triggered=d["e1"], block_pass=d["block_pass"],
            objective_findings=d["objective_findings"], verdict=d["verdict"],
        )
    except (json.JSONDecodeError, KeyError, OSError):
        return None


def _all_checkpoints(patient: str) -> list["SummaryEval"]:
    """Все сохранённые результаты пациента из чекпоинтов (для --report-only:
    пересобрать отчёт без повторного прогона LLM)."""
    ckpt_dir = _ckpt_dir()
    out: list[SummaryEval] = []
    for f in sorted(ckpt_dir.glob(f"{patient}__row*.json")):
        row = int(f.stem.split("row")[-1])
        ev = _load_checkpoint(patient, row)
        if ev is not None:
            out.append(ev)
    return out


def _spec_for_row(row: int) -> tuple[str, str]:
    if row == 1:
        return ("эталон (gold)", "gold")
    spec = ROW_DISTORTIONS.get(row)
    if spec is None:
        return ("(нет в легенде)", "-")
    return (spec.dtype, spec.severity)


def _count_block_pass(block_report: dict, subcriteria: tuple[str, ...]) -> str:
    passed = sum(1 for c in subcriteria
                 if isinstance(block_report.get(c), dict) and block_report[c].get("pass") is True)
    return f"{passed}/{len(subcriteria)}"


def _block_pass_majority(reports: list[dict], block: str, subs: tuple[str, ...]) -> str:
    """Пройденные подкритерии блока по МАЖОРИТАРНОМУ голосу судей (не по одному
    «репрезентативному» отчёту). Подкритерии без валидных голосов (sentinel-блок
    из-за сбоя JSON, `pass=None`) считаются «нет данных» и НЕ засчитываются как
    провал; если по блоку нет данных вовсе — «н/д». Так сбой JSON никогда не
    превращается в фальшивый «0/N» и не обнуляет оценку."""
    passed = nodata = 0
    for code in subs:
        votes = [rep.get(block, {}).get(code, {}).get("pass")
                 for rep in reports if isinstance(rep.get(block, {}).get(code), dict)]
        votes = [v for v in votes if isinstance(v, bool)]
        if not votes:
            nodata += 1
        elif sum(votes) > len(votes) / 2:
            passed += 1
    if nodata == len(subs):
        return "н/д"
    return f"{passed}/{len(subs)}" + (f" (н/д:{nodata})" if nodata else "")


def load_patient(xlsx_path: str, patient: str) -> tuple[str, dict[int, str]]:
    """Возвращает (исходник строки 0, {row: текст суммаризации} для row 1..38)."""
    df = pd.read_excel(xlsx_path, sheet_name=SHEET_NAME)
    if patient not in df.columns:
        raise SystemExit(f"Колонка пациента {patient!r} не найдена. Доступны: "
                         f"{[c for c in df.columns if c != ROW_LABEL_COL]}")
    col = df[patient]
    source = str(col.iloc[0]).strip()
    summaries: dict[int, str] = {}
    for n in range(1, len(col)):
        val = col.iloc[n]
        if isinstance(val, str) and val.strip():
            summaries[n] = val.strip()
    return source, summaries


def evaluate_one(source: str, summary: str, *, row: int, patient: str,
                 panel: JudgePanel, no_preprocessing: bool = False) -> SummaryEval:
    """
    Прогон одной пары (исходник, суммаризация-кандидат) через LLM-судей.

    Обычный режим (no_preprocessing=False):
      1) pre-evaluation gate (scope=None) — фиксируем вердикт и причины;
      2) объективный слой — фактологическая сверка с исходником;
      3) три раунда LLM-судей (R1 независимо -> R2 cross-review -> R3 агрегатор).

    Режим «только LLM-судьи» (no_preprocessing=True, по решению НПКЦ): Блоки 1-3
    ОТКЛЮЧЕНЫ (нет gate, нет объективного слоя, нет сегментации) — судьи получают
    только сырой источник + суммаризацию + таксономию, а заземляющие поля промптов
    нейтральны. Прогоняются только судьи (R1->R2) и финальный арбитр (R3).
    """
    dtype, severity = _spec_for_row(row)
    t0 = time.time()
    log.info("─" * 64)
    log.info("[row %d] %s (%s) | %d симв.%s", row, dtype, severity, len(summary),
             "  [режим: только LLM-судьи]" if no_preprocessing else "")

    if no_preprocessing:
        # Блоки 1-3 отключены — контекст без заземления, gate/объективный слой не считаем.
        gate_status, gate_reasons = "—", []
        ctx = judge.build_context(source, summary, scope=None, panel=panel, grounded=False)
        findings: list[str] = []   # объективный слой не применялся
        log.info("  препроцессинг отключён: gate и объективный слой пропущены")
    else:
        # ── Этап 1: pre-evaluation gate ──────────────────────────────
        gate_decision = gate.evaluate_gate(source, summary, scope=None, panel=panel)
        gate_status, gate_reasons = gate_decision.status, gate_decision.reason_codes()
        log.info("  gate: %s  причины=%s", gate_status, gate_reasons)
        # ── Этап 2: объективный слой + контекст судей ────────────────
        ctx = judge.build_context(source, summary, scope=None, panel=panel,
                                  gate_decision=gate_decision)
        findings = ctx.obj_report.hard_findings()
        log.info("  объективный слой: %d жёстких находок", len(findings))

    # ── Этап 3: три раунда судей ─────────────────────────────────
    roles = list(config.JUDGE_ROLES)
    r1: dict[str, Optional[dict]] = {}
    for role in roles:
        try:
            r1[role] = judge.score_round1(panel, role, ctx)
        except LLMError as e:
            log.error("    %s провал R1: %s", role, e)
            r1[role] = None
    valid = [r for r in roles if r1.get(r) is not None]
    if not valid:
        return SummaryEval(row, dtype, severity, len(summary), gate_status,
                           gate_reasons, "ошибка", False, {}, findings,
                           "R1: все судьи провалились")

    import random
    shuffled = valid.copy()
    random.shuffle(shuffled)
    r2: dict[str, dict] = {}
    for role in shuffled:
        peers = [r for r in shuffled if r != role] or [role]
        while len(peers) < 2:
            peers.append(peers[-1])
        try:
            r2[role] = judge.score_round2(panel, role, ctx, r1[role],
                                          [r1[peers[0]], r1[peers[1]]])
        except LLMError as e:
            log.warning("    %s провал R2 (%s) — беру R1", role, e)
            r2[role] = r1[role]

    r1_list = [r1[r] for r in roles if r1.get(r) is not None]
    r2_list = [r2.get(r, r1[r]) for r in roles if r1.get(r) is not None]
    while len(r1_list) < 3:
        r1_list.append(r1_list[-1]); r2_list.append(r2_list[-1])
    try:
        r3 = judge.score_round3(panel, config.AGGREGATOR_ROLE, ctx, r1_list, r2_list)
    except LLMError as e:
        log.error("    агрегатор провалился: %s", e)
        r3 = {"category": "ошибка", "e1_triggered": False, "verdict": str(e),
              "summary_by_block": {}}

    e1_signals = judge._collect_e1_signals(r1, r2, r3)
    e1 = bool(e1_signals["aggregator_flagged"] or e1_signals["raised_by_judges"])

    # Подсчёт пройденных подкритериев по уточнённым (R2) отчётам — мажоритарно по
    # всем судьям; sentinel-блоки (сбой JSON) считаются «нет данных», не «0/N».
    final_reports = [r2.get(r, r1[r]) for r in roles if r1.get(r) is not None]
    block_pass = {block: _block_pass_majority(final_reports, block, subs)
                  for block, subs in judge.TAXONOMY.items()}

    category = r3.get("category", "—")

    # Для сверки/аудита — детерминированный итог Блока 5 (aggregator.finalize)
    # поверх тех же голосов судей. Первичным считаем арбитра R3; это лишь лог.
    try:
        import aggregator
        det = aggregator.finalize({
            "gate": {"status": None},
            "r1": {r: r1[r] for r in roles if r1.get(r) is not None},
            "r2": r2, "r3": r3, "e1_signals": e1_signals,
        })
        log.info("  [сверка] детерминированный агрегатор (Блок 5): %s", det.get("category"))
    except Exception as agg_err:  # noqa: BLE001 — сверка не должна ронять прогон
        log.debug("  детерминированная сверка недоступна: %s", agg_err)

    log.info("  ИТОГ row %d: %s%s  (%.0f с)", row, category, "  [E1!]" if e1 else "",
             time.time() - t0)

    result = SummaryEval(row, dtype, severity, len(summary), gate_status,
                         gate_reasons, category, e1, block_pass,
                         findings, r3.get("verdict", ""))

    # чекпоинт на диск (переживает перезапуск)
    ckpt = _ckpt_dir() / f"{patient}__row{row}.json"
    ckpt.parent.mkdir(parents=True, exist_ok=True)
    ckpt.write_text(json.dumps({
        "row": row, "dtype": dtype, "severity": severity, "chars": len(summary),
        "gate_status": result.gate_status, "gate_reasons": result.gate_reasons,
        "category": category, "e1": e1, "block_pass": block_pass,
        "objective_findings": findings, "verdict": result.verdict,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def write_report(path: str, patient: str, source: str,
                 results: list[SummaryEval]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    ranked = sorted(results, key=lambda r: r.rank_key(), reverse=True)
    wb = Workbook()
    bold = Font(bold=True)
    gold_fill = PatternFill("solid", fgColor="FFE699")
    crit_fill = PatternFill("solid", fgColor="F8CBAD")
    ben_fill = PatternFill("solid", fgColor="C6E0B4")

    ws = wb.active
    ws.title = "Ранжирование"
    ws.append([f"Пациент {patient} — ранжирование {len(results)} суммаризаций (лучшая сверху)"])
    ws["A1"].font = bold
    ws.append([])
    headers = ["Место", "Строка", "Тип искажения", "Критичность", "Симв.",
               "Gate", "Категория судей", "E1", "A", "B", "C", "D", "E",
               "Реальных ошибок", "Компрессия (пропуски)"]
    ws.append(headers)
    for c in ws[3]:
        c.font = bold
    for place, r in enumerate(ranked, 1):
        ws.append([place, r.row, r.distortion_type, r.severity, r.chars,
                   r.gate_status, r.category, "ДА" if r.e1_triggered else "—",
                   r.block_pass.get("A", ""), r.block_pass.get("B", ""),
                   r.block_pass.get("C", ""), r.block_pass.get("D", ""),
                   r.block_pass.get("E", ""), r.n_real_errors, r.n_compression])
        fill = (gold_fill if r.severity == "gold" else
                crit_fill if r.severity == "critical" else
                ben_fill if r.severity == "benign" else None)
        if fill:
            ws[ws.max_row][2].fill = fill

    ws2 = wb.create_sheet("Объективные находки")
    ws2.append(["Строка", "Тип", "Критичность", "Жёсткие находки объективного слоя"])
    for c in ws2[1]:
        c.font = bold
    for r in sorted(results, key=lambda x: x.row):
        ws2.append([r.row, r.distortion_type, r.severity, " | ".join(r.objective_findings) or "—"])
        ws2[ws2.max_row][3].alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["D"].width = 90

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    log.info("Excel-отчёт сохранён -> %s", path)


def render_console(patient: str, results: list[SummaryEval]) -> str:
    ranked = sorted(results, key=lambda r: r.rank_key(), reverse=True)
    best = ranked[0] if ranked else None
    llm_only = bool(results) and all(r.gate_status == "—" for r in results)
    lines = [
        "=" * 78,
        f"ПАЦИЕНТ {patient} — РЕЗУЛЬТАТ СКВОЗНОГО ПРОГОНА ({len(results)} суммаризаций)"
        + ("  [РЕЖИМ: ТОЛЬКО LLM-СУДЬИ]" if llm_only else ""),
        "=" * 78,
        "",
    ]
    if llm_only:
        lines += ["ПРЕПРОЦЕССИНГ (Блоки 1-3) ОТКЛЮЧЁН: gate и объективный слой не",
                  "применялись; оценивают только LLM-судьи и финальный арбитр.", ""]
    else:
        lines.append("ЭТАП 1 — PRE-EVALUATION GATE:")
        rejected = sum(1 for r in results if r.gate_status == "reject")
        passed = len(results) - rejected
        lines.append(f"  pass={passed}/{len(results)}, reject={rejected}/{len(results)}")
        lines.append("  (scope-aware: легитимные сводки проходят, блокируется лишь полное")
        lines.append("   отсутствие критической категории — диагнозы/препараты; тонкую")
        lines.append("   дискриминацию good/bad дают судьи)")
    lines += ["", ("ЭТАП: LLM-СУДЬИ, РАНЖИРОВАНИЕ (лучшая сверху):" if llm_only
                   else "ЭТАП 2 — LLM-СУДЬИ, РАНЖИРОВАНИЕ (лучшая сверху):")]
    if not llm_only:
        lines.append("  (реал.ошиб = искажения чисел/полярности/причинности/интерпретации; "
                     "компр. = пропущенные сущности = закономерное сжатие, не ошибка)")
    else:
        lines.append("  (объективный слой отключён → реал.ош/компр. не вычисляются, здесь 0)")
    lines.append("")
    lines.append(f"  {'#':<3}{'стр':<5}{'тип искажения':<26}{'крит.':<10}"
                 f"{'категория':<28}{'E1':<4}{'реал.ош':<8}{'компр.'}")
    for place, r in enumerate(ranked, 1):
        lines.append(f"  {place:<3}{r.row:<5}{r.distortion_type[:24]:<26}{r.severity:<10}"
                     f"{r.category[:26]:<28}{'ДА' if r.e1_triggered else '—':<4}"
                     f"{r.n_real_errors:<8}{r.n_compression}")
    if best:
        lines += ["", "─" * 78,
                  f"ЛУЧШАЯ СУММАРИЗАЦИЯ: строка {best.row} — {best.distortion_type} "
                  f"({best.severity})",
                  f"  категория: {best.category} | E1: {'да' if best.e1_triggered else 'нет'} "
                  f"| реальных ошибок: {best.n_real_errors} | компрессия: {best.n_compression}",
                  f"  вердикт: {best.verdict[:300]}"]
        gold = next((r for r in results if r.severity == "gold"), None)
        if gold:
            gold_place = next(i for i, r in enumerate(ranked, 1) if r.row == gold.row)
            verdict = ("✓ алгоритм поставил эталон на 1-е место"
                       if gold_place == 1 else
                       f"⚠ эталон на {gold_place}-м месте (не первый) — разобрать")
            lines.append(f"  Проверка корректности: {verdict}")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--patient", default="Ж1")
    ap.add_argument("--limit-rows", type=int, default=None,
                    help="ограничить число суммаризаций (для пробы)")
    ap.add_argument("--representative", action="store_true",
                    help="эталон (стр.1) + по одному представителю каждого типа "
                         "искажения (~17 строк вместо 38) — быстрее, покрывает все типы")
    ap.add_argument("--report-only", action="store_true",
                    help="пересобрать отчёт из чекпоинтов без прогона LLM")
    ap.add_argument("--recompute-objective", action="store_true",
                    help="пересчитать ТОЛЬКО объективный слой (новым кодом) для уже "
                         "посчитанных чекпоинтов, переиспользуя результаты судей; обновляет "
                         "чекпоинты и отчёт. Дёшево (без раундов судей).")
    ap.add_argument("--no-preprocessing", action="store_true",
                    help="режим «только LLM-судьи»: отключить Блоки 1-3 (препроцессор, "
                         "объективный слой, gate); прогнать только судей (R1->R2) и "
                         "финального арбитра (R3). Чекпоинты/вывод — в отдельном "
                         "пространстве, чтобы не смешивать с grounded-прогоном.")
    ap.add_argument("--profile", default=None)
    ap.add_argument("--xlsx", default=None)
    ap.add_argument("--out", default="reports/eval_patient",
                    help="префикс выходных файлов (.xlsx + .log)")
    args = ap.parse_args()

    # Изолируем чекпоинты режима «только LLM-судьи» от grounded-результатов.
    global CKPT_SUBDIR
    if args.no_preprocessing:
        CKPT_SUBDIR = "eval_patient_llmonly"

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s %(name)s: %(message)s",
        handlers=[logging.FileHandler(f"{args.out}.log", encoding="utf-8"),
                  logging.StreamHandler()],
    )

    xlsx = args.xlsx or next(x for x in glob.glob("../materials/*.xlsx") if "~$" not in x)
    source, summaries = load_patient(xlsx, args.patient)

    if args.report_only:
        # Пересобрать отчёт из уже посчитанных чекпоинтов — без обращения к LLM.
        results = _all_checkpoints(args.patient)
        if not results:
            raise SystemExit(f"Нет чекпоинтов для пациента {args.patient} "
                             f"(checkpoints/{CKPT_SUBDIR}/{args.patient}__row*.json)")
        log.info("report-only: собрано %d результатов из чекпоинтов", len(results))
        print("\n" + render_console(args.patient, results))
        write_report(f"{args.out}.xlsx", args.patient, source, results)
        return

    if args.recompute_objective:
        # Пересчитать объективный слой новым кодом, переиспользуя вердикты судей
        # (дорогую часть) из чекпоинтов. Обновляет objective_findings в чекпоинтах.
        panel = JudgePanel(args.profile)
        objective_layer.clear_entity_cache()
        ckpt_dir = _ckpt_dir()
        results: list[SummaryEval] = []
        for fpath in sorted(ckpt_dir.glob(f"{args.patient}__row*.json")):
            row = int(fpath.stem.split("row")[-1])
            if row not in summaries:
                continue
            d = json.loads(fpath.read_text(encoding="utf-8"))
            rep = objective_layer.compare_texts(source, summaries[row], panel=panel)
            findings = rep.hard_findings()
            d["objective_findings"] = findings
            fpath.write_text(json.dumps(d, ensure_ascii=False, indent=2), encoding="utf-8")
            results.append(SummaryEval(
                row=row, distortion_type=d["dtype"], severity=d["severity"],
                chars=d["chars"], gate_status=d["gate_status"], gate_reasons=d["gate_reasons"],
                category=d["category"], e1_triggered=d["e1"], block_pass=d["block_pass"],
                objective_findings=findings, verdict=d["verdict"]))
            log.info("row %d (%s): пересчитан объективный слой — %d находок "
                     "(реальных ошибок: %d)", row, d["dtype"], len(findings),
                     results[-1].n_real_errors)
        print("\n" + render_console(args.patient, results))
        write_report(f"{args.out}.xlsx", args.patient, source, results)
        return

    if args.representative:
        # эталон (стр.1) + первая (минимальная по номеру) строка каждого типа искажения
        first_of_type: dict[str, int] = {}
        for row, spec in sorted(ROW_DISTORTIONS.items()):
            first_of_type.setdefault(spec.dtype, row)
        rows = sorted({1, *first_of_type.values()} & set(summaries))
    elif args.limit_rows:
        rows = sorted(summaries)[: args.limit_rows]
    else:
        rows = sorted(summaries)
    mode_note = " | режим: ТОЛЬКО LLM-судьи (без препроцессинга)" if args.no_preprocessing else ""
    log.info("Пациент %s: исходник %d симв., %d суммаризаций-кандидатов (строки %s..%s)%s",
             args.patient, len(source), len(rows), rows[0], rows[-1], mode_note)

    panel = JudgePanel(args.profile)
    objective_layer.clear_entity_cache()

    results: list[SummaryEval] = []
    for i, row in enumerate(rows, 1):
        cached = _load_checkpoint(args.patient, row)
        if cached is not None:
            log.info("══ %d/%d (строка %d) — из чекпоинта, пропуск ══", i, len(rows), row)
            results.append(cached)
            continue
        log.info("══ %d/%d (строка %d) ══", i, len(rows), row)
        try:
            results.append(evaluate_one(source, summaries[row], row=row,
                                        patient=args.patient, panel=panel,
                                        no_preprocessing=args.no_preprocessing))
        except Exception as e:  # noqa: BLE001 — длинный прогон не должен падать целиком
            log.exception("row %d: непредвиденная ошибка: %s", row, e)

    print("\n" + render_console(args.patient, results))
    write_report(f"{args.out}.xlsx", args.patient, source, results)
    log.info("Готово. Лог: %s.log | Отчёт: %s.xlsx | Чекпоинты: checkpoints/%s/",
             args.out, args.out, CKPT_SUBDIR)


if __name__ == "__main__":
    main()
