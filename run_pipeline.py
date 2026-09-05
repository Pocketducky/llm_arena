"""
run_pipeline.py — Блок 7: оркестратор полного end-to-end прогона новой системы.

Назначение (план, Блок 7, критерий проверки): «полный end-to-end прогон
новой системы на текущем датасете (26 ЭМК × 6 моделей) с сохранением
аудит-лога и итогового отчёта».

Связывает воедино все блоки 0-7:
  загрузка пар (data/summaries.xlsx)
        │
        ▼
  judge.evaluate_summary   — Блоки 1-4 (препроцессор, объективный слой,
        │                     шлюз, LLM-as-Judge); чекпоинты на диск
        ▼
  aggregator.finalize      — Блок 5 (детерминированное решение)
        │
        ├──▶ audit.AuditLogger.log_evaluation   — Блок 7 (audit-лог JSONL,
        │                                          версии промптов/моделей/
        │                                          порогов, хэши вместо текста)
        │
        ▼ (после всех пар)
  report.save_results      — Блок 5/7 (Excel: сводная + объективный слой
        │                     рядом с LLM-оценками + статистика + критические)
        ▼
  drift.compute_snapshot   — Блок 7 (снимок распределений для мониторинга
                              дрейфа; сохраняется в audit_log/drift_snapshots/)

ВАЖНО — честно о масштабе полного прогона: 26 ЭМК × 6 моделей = 156 пар,
каждая — это gate (несколько LLM-вызовов извлечения сущностей) + 3 раунда
LLM-as-Judge (раунд 1: 3 судьи × 5 блоков = 15 вызовов; раунд 2: ещё ~3;
раунд 3: 1) — порядка 20+ обращений к Ollama НА ОДНУ ПАРУ. На пилотном
железе (один qwen3:8b на все роли, без параллелизма GPU) это те же часы
на сотни вызовов, что и в Блоке 6 сделали невозможным полный 3-раундовый
прогон на 370 синтетических парах за разумное время сессии. Поэтому здесь:
  • оркестратор написан и обкатан на ВСЮ цепочку (а не на заглушках —
    см. _self_check_dry_run, прогоняющий стаб-evaluation через все слои);
  • демонстрационный прогон --limit делает РЕАЛЬНЫЕ вызовы LLM на
    небольшом срезе пар — доказательство, что проводка «judge ->
    aggregator -> audit -> report -> drift» работает на живых данных
    end-to-end (не только в теории на стабах);
  • полный прогон всех 156 пар — отдельная, многочасовая задача,
    которую целесообразно ставить на ночь/на отдельном железе, и
    которая уже полностью поддерживается этим же скриптом без правок
    кода: `python run_pipeline.py` (без --limit) с уже встроенными
    чекпоинтами judge.py (переживают перезапуск процесса).
"""

from __future__ import annotations

import argparse
import logging
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Optional

import pandas as pd

import aggregator
import audit
import config
import gate
import llm_client
import drift
import judge
import report
from llm_client import JudgePanel

log = logging.getLogger("run_pipeline")

DATA_PATH = Path("data") / "summaries.xlsx"
REPORTS_DIR = Path("reports")


def load_dataset(path: Path = DATA_PATH) -> pd.DataFrame:
    df = pd.read_excel(path, dtype=str).fillna("")
    required = {"emr_id", "model_id", "source_text", "summary_text"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"В {path} не хватает колонок: {sorted(missing)}")
    return df


def _probe_max_model_len() -> object:
    """Длина контекста, с которой реально подняты модели vLLM.

    Клиент об этом не знал вовсе: num_ctx для vLLM не передаётся (окно задаётся
    сервером через --max-model-len), а check_environment делал лишь 200-токенный
    roundtrip. Переполнение проявлялось как HTTP 400 уже посреди многочасового
    прогона. Значение отдаёт сам сервер в GET /v1/models.
    """
    if config.LLM_BACKEND != "vllm":
        return f"n/a (бэкенд {config.LLM_BACKEND})"
    try:
        import requests
        lens = {}
        for endpoint in config.vllm_endpoints():
            r = requests.get(f"{endpoint.rstrip('/')}/models",
                             headers={"Authorization": f"Bearer {config.VLLM_API_KEY}"},
                             timeout=5)
            r.raise_for_status()
            for m in r.json().get("data", []):
                if m.get("max_model_len"):
                    lens[m.get("id")] = m["max_model_len"]
        return ", ".join(f"{k}: {v}" for k, v in lens.items()) or "не сообщается сервером"
    except Exception as exc:   # noqa: BLE001 — диагностика не должна ронять отчёт
        return f"не удалось определить ({type(exc).__name__})"


def _log_progress(done: int, total: int, t0: float) -> None:
    """Прогресс с оценкой оставшегося времени: на прогоне в несколько часов
    оператору нужно понимать, укладывается ли он в окно."""
    elapsed = time.monotonic() - t0
    if not done:
        return
    eta = elapsed / done * (total - done)
    log.info("    [%d/%d] прошло %s, осталось ~%s",
             done, total, _hms(elapsed), _hms(eta))


def _hms(seconds: float) -> str:
    seconds = int(max(0, seconds))
    h, rem = divmod(seconds, 3600)
    m, sec = divmod(rem, 60)
    return f"{h}ч {m:02d}м" if h else f"{m}м {sec:02d}с"


def run(*, limit: Optional[int] = None, profile: Optional[str] = None,
        run_id: Optional[str] = None, use_checkpoints: bool = True,
        scope: Optional[str] = config.DATASET_SCOPE,
        concurrency: int = config.CONCURRENCY) -> dict:
    """
    Полный (или усечённый --limit) прогон. Возвращает сводку:
    {n_pairs, results, audit_path, report_path, snapshot_path}.

    scope — декларация ЗАДАЧИ суммаризации в этом датасете (см. подробное
            объяснение в config.DATASET_SCOPE и judge._SCOPE_DESCRIPTIONS).
            По умолчанию берётся из config.DATASET_SCOPE ("radiologist" —
            нынешний датасет содержит ИМЕННО целевые выжимки «для
            рентгенолога», а не полные суммаризации ЭМК). Передайте
            scope=None, только если оцениваете ДРУГОЙ датасет с полными
            суммаризациями — иначе шлюз будет сравнивать целевую выжимку
            со ВСЕЙ картой и систематически отклонять корректные ответы
            как «неполные» (см. gate.evaluate_gate, комментарий про
            «системный ложный reject на целевых выжимках»).
    """
    df = load_dataset()
    if limit is not None:
        df = df.head(limit)
    log.info("Загружено пар: %d (emr_id × model_id), профиль=%s, scope=%s",
             len(df), profile or config.ACTIVE_PROFILE, scope)

    panel = JudgePanel(profile)
    logger = audit.AuditLogger(run_id=run_id)
    log.info("Audit-лог: %s", logger.path)

    results: list[dict] = []
    t0 = time.monotonic()

    def _evaluate_pair(i: int, row) -> dict:
        """Полная обработка одной пары. Вынесена из цикла ради двух вещей:
        телеметрии LLM-вызовов (собирается по паре) и параллельного прогона."""
        emr_id, model_id = row.emr_id, row.model_id
        source_text, summary_text = row.source_text, row.summary_text
        pair_t0 = time.monotonic()

        with llm_client.collect_telemetry() as stats:
            try:
                cached = judge.load_checkpoint(emr_id, model_id) if use_checkpoints else None
                # Принимаем чекпоинт ТОЛЬКО если: (а) он уже содержит "objective"
                # (запись формата ДО Блока 7 честно пересчитывается заново — иначе
                # в audit-логе/отчёте появится молчаливый пробел в новой колонке
                # для части пар, хуже, чем потратить время на пересчёт), И
                # (б) он посчитан с ТЕМ ЖЕ scope, что и текущий прогон — иначе
                # словим ровно ту стэйл-проблему, что обнаружилась эмпирически:
                # запись "EMR_10__2", посчитанная со scope=None (gate=reject —
                # ложный отказ из-за сравнения целевой выжимки со ВСЕЙ картой),
                # молча подставлялась бы дальше даже после исправления scope.
                cached_ok = (cached is not None and "objective" in cached
                             and cached.get("scope") == scope)
                if cached_ok:
                    log.info("[%d/%d] %s / %s — из чекпоинта (scope=%s)",
                             i, len(df), emr_id, model_id, scope)
                    evaluation = cached
                else:
                    if cached is not None and not cached_ok:
                        log.info("[%d/%d] %s / %s — чекпоинт устарел (scope: %s -> %s), пересчёт...",
                                 i, len(df), emr_id, model_id, cached.get("scope"), scope)
                    else:
                        log.info("[%d/%d] %s / %s — оценка...", i, len(df), emr_id, model_id)
                    evaluation = judge.evaluate_summary(source_text, summary_text, emr_id, model_id,
                                                         scope=scope, panel=panel)
                    if use_checkpoints:
                        judge.save_checkpoint(evaluation)
            except Exception as exc:   # noqa: BLE001
                # Один сбойный ответ модели не должен останавливать многочасовой
                # прогон. ВАЖНО: category="ошибка" отсюда БОЛЬШЕ НЕ перезаписывается
                # агрегатором на «Неприемлемо» (см. aggregator.finalize) — отказ
                # инфраструктуры не выдаётся за клинический вердикт.
                log.error("[%d/%d] %s / %s — СБОЙ: %s — пара помечена «ошибка», прогон продолжается",
                          i, len(df), emr_id, model_id, exc)
                stats.note(f"исключение: {type(exc).__name__}: {exc}"[:300])
                evaluation = judge._err(emr_id, model_id, f"Прогон прерван исключением: {exc}")

            try:
                finalized = aggregator.finalize(evaluation)
            except Exception as agg_exc:   # noqa: BLE001 — агрегация не должна ронять прогон
                log.exception("[%d/%d] %s / %s — сбой агрегации: %s",
                              i, len(df), emr_id, model_id, agg_exc)
                stats.note(f"сбой агрегации: {agg_exc}"[:300])
                finalized = {**judge._err(emr_id, model_id,
                                          f"Сбой детерминированной агрегации: {agg_exc}"),
                             "blocks": {}, "decision_path": [f"Агрегация упала: {agg_exc}"],
                             "e1_triggered": False, "e1_sources": []}

        # Телеметрия (обрывы, ретраи, таймауты, токены, латентность) раньше
        # существовала только как строки в stdout: по итоговому Excel нельзя было
        # отличить «судьи нашли проблемы» от «мы не смогли разобрать ответ судьи».
        finalized["telemetry"] = {**stats.as_dict(),
                                  "wall_seconds": round(time.monotonic() - pair_t0, 1)}
        finalized.setdefault("scope", scope)

        logger.log_evaluation(emr_id=emr_id, model_id=model_id, evaluation=finalized,
                              source_text=source_text, summary_text=summary_text,
                              profile_name=profile, scope=scope)
        return finalized

    rows = list(df.itertuples(index=False))
    workers = max(1, int(concurrency))
    if workers > 1:
        log.info("Параллельный прогон: %d пар одновременно "
                 "(внутри пары порядок R1->R2->R3 сохраняется)", workers)
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = {pool.submit(_evaluate_pair, i, row): i
                       for i, row in enumerate(rows, 1)}
            done: dict[int, dict] = {}
            for fut in as_completed(futures):
                idx = futures[fut]
                done[idx] = fut.result()
                _log_progress(len(done), len(rows), t0)
        results = [done[i] for i in sorted(done)]     # порядок корпуса, а не завершения
    else:
        for i, row in enumerate(rows, 1):
            finalized = _evaluate_pair(i, row)
            results.append(finalized)
            _log_progress(i, len(rows), t0)

    for finalized in results:
        log.info("    -> %s (шлюз=%s, объективный=%s)", finalized.get("category"),
                 (finalized.get("gate") or {}).get("status"),
                 "есть" if finalized.get("objective") else "—")

    elapsed = time.monotonic() - t0
    log.info("Готово: %d пар за %.1f сек (%.1f сек/пара)", len(results), elapsed,
             elapsed / len(results) if results else 0.0)

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    report_path = REPORTS_DIR / f"{logger.run_id}.xlsx"
    run_meta = {
        "run_id": logger.run_id,
        "profile": profile or config.ACTIVE_PROFILE,
        "scope": scope,
        "concurrency": workers,
        "n_pairs": len(results),
        "dataset": str(DATA_PATH),
        "elapsed": _hms(elapsed),
        "max_model_len": _probe_max_model_len(),
    }
    report.save_results(results, str(report_path), run_meta=run_meta)
    log.info("Отчёт сохранён: %s", report_path)

    snapshot_path = None
    if results:
        snap = drift.compute_snapshot(audit.load_entries(logger.path), run_id=logger.run_id, profile=profile)
        snapshot_path = snap.save()
        log.info("Снимок дрейфа сохранён: %s", snapshot_path)
        log.info("\n%s", snap.render())

    return {"n_pairs": len(results), "results": results, "audit_path": logger.path,
            "report_path": report_path, "snapshot_path": snapshot_path}


# ══════════════════════════════════════════════════════════════════
# СУХОЙ ПРОГОН — проверяет проводку judge-результат -> aggregator ->
# audit -> report -> drift БЕЗ единого вызова LLM (на стаб-evaluation,
# повторяющих реальную форму evaluate_summary для pass/reject путей).
# Цель — отделить «правильно ли соединены модули» (проверяется здесь,
# мгновенно и детерминированно) от «что говорит конкретная LLM на
# конкретных данных» (проверяется --limit-прогоном на живых вызовах).
# ══════════════════════════════════════════════════════════════════

def _judge_report(*, failed: tuple = (), nodata: tuple = ()) -> dict:
    """Полный отчёт судьи A-E для стабов: по умолчанию всё пройдено."""
    import judge as _j
    rep: dict = {}
    for block, codes in _j.TAXONOMY.items():
        if block in nodata:
            rep[block] = {c: {"pass": None, "comment": "нет данных: сбой JSON"} for c in codes}
            rep[block][_j.PARSE_ERROR_KEY] = True
            continue
        rep[block] = {c: {"pass": block not in failed or c != codes[0], "comment": "ок"}
                      for c in codes}
    rep["A"].update({"hallucinations": [], "wrong_values": []})
    for c in _j.TAXONOMY["B"]:
        if isinstance(rep["B"].get(c), dict):
            rep["B"][c].setdefault("missing", [])
    if isinstance(rep["E"].get("E1"), dict):
        rep["E"]["E1"].setdefault("danger_examples", [])
    return rep


def _stub_evaluation(emr_id: str, model_id: str, *, kind: str) -> dict:
    """Стабы результата judge.evaluate_summary для проверки проводки без LLM.

    Набор видов отражает семантику v2:
      "pass"        — шлюз pass, судьи всё подтвердили;
      "gate_reject" — шлюз отклонил, НО судьи отработали. В v1 такой пары не
                      существовало (шлюз отсекал её до судей и она получала
                      «Неприемлемо» без свидетельств); теперь это штатный путь,
                      и правильный исход — «Требует редактирования»;
      "incomplete"  — ответы судей не разобраны по трём блокам: это сбой
                      разбора, а не клинический вердикт;
      "error"       — прогон пары упал на исключении.
    """
    base = {"emr_id": emr_id, "model_id": model_id, "scope": "radiologist",
            "e1_signals": {"raised_by_judges": [], "aggregator_flagged": False,
                           "aggregator_named": [], "aggregator_category": None,
                           "consistent": True},
            "objective": {
                "numeric": {"total_in_a": 8, "matched": 8, "mismatch_count": 0,
                            "unit_mismatch_count": 0, "mismatches": [], "unit_mismatches": []},
                "polarity": {"total_in_a": 3, "matched": 3, "flip_count": 0, "flips": []},
                "entities": None,
            }}

    if kind == "error":
        import judge as _j
        return _j._err(emr_id, model_id, "Прогон прерван исключением: ConnectionError")

    if kind == "incomplete":
        rep = _judge_report(nodata=("A", "B", "C"))
        r1 = {"judge_1": rep, "judge_2": rep, "judge_3": rep}
        return {**base, "category": "—", "verdict": "",
                "gate": {"status": "pass", "reasons": [], "coverage": {"declared_scope": "radiologist"}},
                "r1": r1, "r2": r1, "r3": {"category": "Требует редактирования", "verdict": "ок"}}

    if kind == "gate_reject":
        rep = _judge_report()
        r1 = {"judge_1": rep, "judge_2": rep, "judge_3": rep}
        return {**base, "category": "—", "verdict": "",
                "objective": None,
                "gate": {"status": "reject",
                         "reasons": ["entity_recall_low:diagnoses"],
                         "messages": ["низкая полнота сущностей категории «diagnoses»"],
                         "degenerate": False,
                         "coverage": {"declared_scope": "radiologist"}},
                "r1": r1, "r2": r1,
                "r3": {"category": "Требует редактирования", "verdict": "ок"}}

    rep = _judge_report()
    r1 = {"judge_1": rep, "judge_2": rep, "judge_3": rep}
    return {**base,
            "category": "Готово к клиническому применению",
            "verdict": "Сводка точна и полна.",
            "summary_by_block": {b: "ок" for b in "ABCDE"},
            "gate": {"status": "pass", "reasons": [], "coverage": {"declared_scope": "radiologist"}},
            "r1": r1, "r2": r1,
            "r3": {"category": "Готово к клиническому применению", "verdict": "ок",
                   "summary_by_block": {}}}


def _self_check_dry_run():
    """
    Сухой прогон проводки (без LLM): берём по одной паре каждого вида
    исхода (pass / gate-reject), пропускаем через aggregator.finalize ->
    audit.AuditLogger -> report.save_results -> drift.compute_snapshot
    и проверяем, что результат на каждом стыке выглядит так, как должен.
    """
    import tempfile
    import shutil

    tmpdir = Path(tempfile.mkdtemp(prefix="run_pipeline_selfcheck_"))
    try:
        finalized = [
            aggregator.finalize(_stub_evaluation("EMR_DRY1", "stub-model-A", kind="pass")),
            aggregator.finalize(_stub_evaluation("EMR_DRY2", "stub-model-A", kind="gate_reject")),
            aggregator.finalize(_stub_evaluation("EMR_DRY1", "stub-model-B", kind="incomplete")),
            aggregator.finalize(_stub_evaluation("EMR_DRY2", "stub-model-B", kind="error")),
        ]
        got = [f["category"] for f in finalized]
        expected = [aggregator.CATEGORY_READY,      # шлюз pass + судьи всё подтвердили
                    aggregator.CATEGORY_EDIT,       # шлюз reject НЕ даёт «Готово» (баг v1)
                    aggregator.CATEGORY_INCOMPLETE, # сбой разбора != клинический отказ
                    aggregator.CATEGORY_ERROR]      # исключение != клинический отказ
        assert got == expected, f"категории: {got}, ожидалось {expected}"

        logger = audit.AuditLogger(run_id="dry-run-selfcheck", directory=tmpdir / "audit_log")
        for f in finalized:
            logger.log_evaluation(emr_id=f["emr_id"], model_id=f["model_id"], evaluation=f,
                                  source_text=f"исходник {f['emr_id']}", summary_text=f"сводка {f['model_id']}")
        assert logger.count == 4
        entries = audit.load_entries(logger.path)
        assert len(entries) == 4
        assert entries[0]["objective"]["numeric"]["matched"] == 8
        assert entries[1]["objective"] is None
        assert entries[0]["scope"] == "radiologist", "scope снова теряется в аудите"

        report_path = tmpdir / "dry_run_report.xlsx"
        report.save_results(finalized, str(report_path))
        assert report_path.exists()
        import openpyxl
        wb = openpyxl.load_workbook(report_path)
        assert "Сводная таблица" in wb.sheetnames
        ws = wb["Сводная таблица"]
        hdr = [c.value for c in ws[1]]
        assert "Объективный слой\n(Блок 2)" in hdr
        col = hdr.index("Объективный слой\n(Блок 2)") + 1
        cell_vals = [ws.cell(row=r, column=col).value for r in (2, 3, 4)]
        assert any("расхождений не найдено" in (v or "") for v in cell_vals)
        assert any("шлюз отклонил" in (v or "") for v in cell_vals)

        snap = drift.compute_snapshot(entries, run_id="dry-run-selfcheck")
        assert snap.n == 4
        # Ни одна из четырёх пар не должна получить «Неприемлемо»: клинический
        # отказ выносится только за клинические основания.
        assert snap.category_rates.get("Неприемлемо", 0) == 0, snap.category_rates
        snap_path = snap.save(directory=tmpdir / "snapshots")
        assert snap_path.exists()

        print("OK: судья-стаб -> aggregator.finalize -> audit -> report -> drift — проводка цела")
        print(f"  audit: {logger.path.name} ({logger.count} записей)")
        print(f"  report: {report_path.name} (лист 'Сводная таблица', колонка "
              f"'Объективный слой' на месте)")
        print(f"  drift snapshot: {snap_path.name} (n={snap.n}, "
              f"reject_rate={snap.category_rates.get('Неприемлемо', 0):.1%})")
        print("Самопроверка run_pipeline.py (сухой прогон проводки) — OK")
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def _build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--limit", type=int, default=None,
                   help="оценить только первые N пар датасета (демонстрационный прогон; "
                        "без флага — ВСЕ 156 пар, часы на пилотном железе)")
    p.add_argument("--profile", default=None, help="имя профиля моделей (по умолчанию — config.ACTIVE_PROFILE)")
    p.add_argument("--run-id", default=None, help="идентификатор прогона (по умолчанию — метка времени)")
    p.add_argument("--concurrency", type=int, default=config.CONCURRENCY,
                   help="сколько пар обрабатывать параллельно (по умолчанию %(default)s). "
                        "1 — последовательно, как раньше. Внутри пары порядок "
                        "R1->R2->R3 сохраняется в любом случае.")
    p.add_argument("--no-checkpoints", action="store_true", help="игнорировать сохранённые чекпоинты judge.py")
    p.add_argument("--scope", default=None,
                   help="декларация задачи суммаризации (по умолчанию — config.DATASET_SCOPE, "
                        "сейчас 'radiologist' — целевые выжимки для рентгенолога). "
                        "Передайте 'none', если оцениваете датасет с ПОЛНЫМИ суммаризациями ЭМК "
                        "(иначе шлюз будет систематически отклонять корректные целевые выжимки "
                        "как «неполные» — см. config.DATASET_SCOPE)")
    p.add_argument("--dry-run", action="store_true",
                   help="не вызывать LLM вовсе — прогнать самопроверку проводки на стаб-данных")
    return p


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(name)s: %(message)s")
    args = _build_argparser().parse_args()

    if args.dry_run:
        _self_check_dry_run()
        sys.exit(0)

    if args.scope is None:
        scope = config.DATASET_SCOPE
    elif args.scope.strip().lower() in ("none", "null", "-", ""):
        scope = None
    else:
        scope = args.scope
    gate.validate_scope(scope)   # опечатка в --scope включила бы строгий порог без фильтра

    summary = run(limit=args.limit, profile=args.profile, run_id=args.run_id,
                  use_checkpoints=not args.no_checkpoints, scope=scope,
                  concurrency=args.concurrency)
    print()
    print(f"Готово: {summary['n_pairs']} пар")
    print(f"  audit-лог: {summary['audit_path']}")
    print(f"  отчёт:     {summary['report_path']}")
    print(f"  снимок:    {summary['snapshot_path']}")
