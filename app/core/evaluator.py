"""
Арена LLM v3.1 — Оценка суммаризаций медицинских ЭМК
Архитектура: R1 (3 судьи) → R2 (каждый читает двух других) → R3 (финальный арбитр)

Модели — все быстрые, без thinking-режима:
  LLM_A: llama3.1:8b    — 128k контекст, Meta
  LLM_B: mistral:7b     — стабильный JSON, хороший русский, 4.1GB
  LLM_C: qwen2.5:7b     — 128k контекст, лучший русский среди 7B

Скорость: ~8 мин на суммаризацию → ~24 часа на 182 шт
  mistral:7b — обычная модель, стабильно возвращает JSON
"""

import json
import re
import time
import random
import logging
from pathlib import Path
from typing import Optional

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import prompts
from app.core.utils import extract_json, _warn_if_all_zeros, _make_minimal_prompt, split_source

# ══════════════════════════════════════════════════════════════════
# КОНФИГУРАЦИЯ
# ══════════════════════════════════════════════════════════════════

MODELS = {
    "LLM_A": {"name": "llama3.1:8b"},
    "LLM_B": {"name": "mistral:7b"},
    "LLM_C": {"name": "qwen2.5:7b"},
}
# Финальный арбитр R3 — mistral:7b
ARBITER = "LLM_B"   # mistral:7b — стабильный арбитр

OLLAMA_URL   = "http://localhost:11434/api/generate"
NUM_CTX      = 16384   # КРИТИЧНО: без этого Ollama дефолтит на 2048
TIMEOUT      = 300     # 5 мин — достаточно для 7B моделей
OLLAMA_SLEEP = 2
MAX_RETRIES  = 2

INPUT_XLSX  = "data/result_data/summaries.xlsx"
OUTPUT_XLSX = "data/results_data/evaluation_results.xlsx"

# Файловый лог — DEBUG (все сырые ответы моделей)
# Консоль — INFO (только прогресс)
_file_handler   = logging.FileHandler("evaluation.log", encoding="utf-8")
_file_handler.setLevel(logging.DEBUG)
_file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))

_console_handler = logging.StreamHandler()
_console_handler.setLevel(logging.INFO)
_console_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))

log = logging.getLogger(__name__)
log.setLevel(logging.DEBUG)
log.addHandler(_file_handler)
log.addHandler(_console_handler)

# ══════════════════════════════════════════════════════════════════
# OLLAMA КЛИЕНТ
# ══════════════════════════════════════════════════════════════════

def ollama_call(model_name: str, prompt: str, force_json: bool = True) -> str:
    """
    force_json=True: добавляет "format":"json" в запрос.
    Это заставляет Ollama генерировать ТОЛЬКО валидный JSON на уровне
    токенов — модель физически не может написать текст вне JSON.
    Отключаем только для retry-промптов где просим исправить текст.
    """
    payload = {
        "model":  model_name,
        "prompt": prompt,
        "stream": False,
        "options": {
            "temperature": 0.1,
            "num_predict": 1024,
            "num_ctx":     NUM_CTX,
        },
    }
    if force_json:
        payload["format"] = "json"

    for attempt in range(1, MAX_RETRIES + 2):
        try:
            r = requests.post(OLLAMA_URL, json=payload, timeout=TIMEOUT)
            r.raise_for_status()
            resp = r.json()
            used = resp.get("prompt_eval_count", "?")
            log.info(f"        токенов: {used}/{NUM_CTX}"
                     + (" [JSON mode]" if force_json else ""))
            return resp["response"]
        except requests.exceptions.Timeout:
            log.warning(f"        таймаут попытка {attempt}/{MAX_RETRIES+1}")
            if attempt <= MAX_RETRIES:
                time.sleep(10)
            else:
                raise

def ask(model_key: str, prompt: str, desc: str = "") -> dict:
    """Запрос к модели с retry при JSON-ошибке.
    Вызывает модель и возвращает спарсенный JSON.
    отключает режим размышлений, скорость как у обычной модели.
    """
    name = MODELS[model_key]["name"]

    # /no_think убран — он конфликтует с format=json и вызывает возврат нулей
    actual_prompt = prompt
    log.info(f"      {model_key}({name}) {desc} | {len(actual_prompt)} симв.")
    last_raw = ""

    for attempt in range(1, 4):
        try:
            if attempt == 1:
                # Первая попытка: основной промпт с format=json
                current = actual_prompt
                fj = True
            elif attempt == 2:
                # Вторая попытка: без format=json
                # отвечает без жёсткого JSON-режима
                log.warning(f"      {model_key} retry2 (без format=json)...")
                current = actual_prompt
                fj = False
            else:
                # Третья попытка: минимальный промпт
                log.warning(f"      {model_key} retry3 (минимальный промпт)...")
                current = _make_minimal_prompt(desc, last_raw)
                fj = True

            raw = ollama_call(name, current, force_json=fj)
            last_raw = raw
            log.debug(f"      RAW попытка{attempt}: {raw[:300]}")

            result = extract_json(raw)
            _warn_if_all_zeros(model_key, result, raw)
            time.sleep(OLLAMA_SLEEP)
            if attempt > 1:
                log.info(f"      {model_key} ✅ получен на попытке {attempt}")
            return result

        except (json.JSONDecodeError, ValueError) as e:
            log.warning(f"      {model_key} попытка{attempt}: {str(e)[:120]}")
            if attempt < 3:
                time.sleep(3)

    raise ValueError(f"{model_key} не вернул JSON после 3 попыток")

# ══════════════════════════════════════════════════════════════════
# ОЦЕНКА ОДНОЙ СУММАРИЗАЦИИ
# ══════════════════════════════════════════════════════════════════

def score_r1(model_key: str, source: str, summary: str) -> dict:
    """R1: три коротких запроса — клиника, инструментальные, штрафы."""
    src_clinical, src_labs = split_source(source)

    clinical     = ask(model_key, prompts.PROMPT_CLINICAL.format(
                       source=src_clinical, summary=summary,
                       fewshot_good_summary=prompts.FEWSHOT_GOOD_SUMMARY,
                       fewshot_good_clinical=prompts.FEWSHOT_GOOD_CLINICAL,
                       fewshot_bad_summary=prompts.FEWSHOT_BAD_SUMMARY,
                       fewshot_bad_clinical=prompts.FEWSHOT_BAD_CLINICAL,
                   ), "клиника")
    instrumental = ask(model_key, prompts.PROMPT_INSTRUMENTAL.format(
                       source=src_labs, summary=summary), "лаб+инстр")
    penalties    = ask(model_key, prompts.PROMPT_PENALTIES.format(
                       source=source, summary=summary,
                       fewshot_good_penalties=prompts.FEWSHOT_GOOD_PENALTIES,
                       fewshot_bad_penalties=prompts.FEWSHOT_BAD_PENALTIES,
                   ), "штрафы")

    comp = float(clinical.get("complaints",      {}).get("score", 0))
    dh   = float(clinical.get("disease_history", {}).get("score", 0))
    co   = float(clinical.get("comorbidities",   {}).get("score", 0))
    hab  = float(clinical.get("habits",          {}).get("score", 0))
    lab  = float(instrumental.get("labs",        {}).get("score", 0))
    img  = float(instrumental.get("imaging",     {}).get("score", 0))
    pen  = float(penalties.get("penalties", 0))

    positive    = comp + dh + co + hab + lab + img
    final_score = max(0.0, min(100.0, round(positive + pen, 1)))

    return {
        "complaints":      comp,
        "disease_history": dh,
        "comorbidities":   co,
        "habits":          hab,
        "labs":            lab,
        "imaging":         img,
        "penalties":       pen,
        "final_score":     final_score,
        "iodine_flag":     bool(penalties.get("iodine_missing", False)),
        "safety_flag":     bool(penalties.get("safety_flag", False)),
        "safety_reason":   str(penalties.get("safety_reason", "")),
        "hallucinations":  list(penalties.get("hallucinations", [])),
        "wrong_values":    list(penalties.get("wrong_values", [])),
        "missing_clinical": list(clinical.get("comorbidities", {}).get("missing", [])),
    }


def score_r2(model_key: str, summary: str,
             my_r1: dict, peer1_r1: dict, peer2_r1: dict) -> dict:
    """R2: пересмотр своей оценки с учётом двух коллег."""

    def compact(r: dict) -> str:
        return json.dumps({k: v for k, v in r.items()
                           if k not in ("missing_clinical","safety_reason",
                                        "wrong_values","hallucinations")},
                          ensure_ascii=False)

    result = ask(model_key, prompts.PROMPT_R2.format(
        summary    = summary,
        my_report  = compact(my_r1),
        peer_1     = compact(peer1_r1),
        peer_2     = compact(peer2_r1),
    ), "R2-пересмотр")

    comp = float(result.get("complaints",      0))
    dh   = float(result.get("disease_history", 0))
    co   = float(result.get("comorbidities",   0))
    hab  = float(result.get("habits",          0))
    lab  = float(result.get("labs",            0))
    img  = float(result.get("imaging",         0))
    pen  = float(result.get("penalties",       0))

    positive    = comp + dh + co + hab + lab + img
    final_score = max(0.0, min(100.0, round(positive + pen, 1)))

    return {
        "complaints":      comp,
        "disease_history": dh,
        "comorbidities":   co,
        "habits":          hab,
        "labs":            lab,
        "imaging":         img,
        "penalties":       pen,
        "final_score":     final_score,
        "iodine_flag":     bool(result.get("iodine_flag", False)),
        "safety_flag":     bool(result.get("safety_flag", False)),
        "hallucinations":  list(result.get("hallucinations", [])),
        "quality":         str(result.get("quality", "—")),
    }


def evaluate_one(source: str, summary: str, emr_id: str, model_id: str) -> dict:
    """
    Полный цикл:
      R1 — три модели оценивают независимо (3 запроса каждая)
      R2 — каждая пересматривает с учётом двух других (1 запрос каждая)
      R3 — арбитр выносит финальный вердикт (1 запрос)
    """
    log.info(f"\n{'═'*60}")
    log.info(f"  ЭМК: {emr_id} | Модель: {model_id}")
    log.info(f"{'═'*60}")

    mk_list = list(MODELS.keys())

    # ── РАУНД 1 ───────────────────────────────────────────────────
    log.info("  [R1] Независимая оценка...")
    r1: dict[str, dict] = {}
    for mk in mk_list:
        try:
            res = score_r1(mk, source, summary)
            r1[mk] = res
            log.info(f"    {mk}: {res['final_score']:.1f}/100"
                     + (" ⚠йод" if res["iodine_flag"] else "")
                     + (" 🚨" if res["safety_flag"] else ""))
        except Exception as e:
            log.error(f"    {mk} R1 провал: {type(e).__name__}: {str(e)[:150]}")
            r1[mk] = {"final_score": None, "error": str(e)}

    valid_r1 = {mk: v for mk, v in r1.items() if v.get("final_score") is not None}
    if not valid_r1:
        log.error("  ❌ R1: все модели провалились")
        return _err(emr_id, model_id, "R1 полный провал")

    # ── РАУНД 2 ───────────────────────────────────────────────────
    log.info("  [R2] Пересмотр с учётом коллег...")
    r2: dict[str, dict] = {}

    # Порядок перемешиваем — против position bias
    shuffled = mk_list.copy()
    random.shuffle(shuffled)

    for mk in shuffled:
        if mk not in valid_r1:
            continue
        others = [k for k in shuffled if k != mk and k in valid_r1]
        # Если меньше двух валидных коллег — дублируем
        while len(others) < 2:
            others.append(others[-1] if others else mk)
        try:
            res = score_r2(mk, summary, valid_r1[mk],
                           valid_r1[others[0]], valid_r1[others[1]])
            r2[mk] = res
            log.info(f"    {mk}: {res['final_score']:.1f}/100 [{res['quality']}]"
                     + (" ⚠йод" if res["iodine_flag"] else "")
                     + (" 🚨" if res["safety_flag"] else ""))
        except Exception as e:
            log.error(f"    {mk} R2 провал: {type(e).__name__}: {str(e)[:150]}")
            # Fallback: оставляем R1
            r2[mk] = {**valid_r1[mk], "quality": "—"}

    # ── РАУНД 3: финальный арбитраж ───────────────────────────────
    log.info("  [R3] Финальный арбитраж...")

    def fmt_r2(mk: str) -> str:
        d = r2.get(mk, r1.get(mk, {}))
        return json.dumps({k: v for k, v in d.items()
                           if k not in ("missing_clinical","safety_reason",
                                        "wrong_values","error")},
                          ensure_ascii=False)

    # Пробуем арбитра, при провале берём любую доступную модель
    r3 = {}
    # Выбираем арбитра: предпочитаем того у кого в R2 ненулевые баллы
    def model_r2_score(mk):
        return r2.get(mk, {}).get("final_score") or 0.0

    arbiter_order = sorted(
        [ARBITER] + [mk for mk in mk_list if mk != ARBITER],
        key=lambda mk: (model_r2_score(mk) > 0, mk == ARBITER),
        reverse=True
    )
    log.info(f"    Порядок арбитров: {arbiter_order} "
             f"(R2 scores: {[(mk, model_r2_score(mk)) for mk in arbiter_order]})")

    for arb_mk in arbiter_order:
        r2_score = model_r2_score(arb_mk)
        if r2_score == 0.0 and len(arbiter_order) > 1:
            log.warning(f"    Пропускаю {arb_mk} как арбитра — R2 score=0")
            continue
        try:
            r3 = ask(arb_mk, prompts.PROMPT_R3.format(
                report_a = fmt_r2(mk_list[0]),
                report_b = fmt_r2(mk_list[1]),
                report_c = fmt_r2(mk_list[2]),
            ), "R3-арбитраж")
            final_r3 = r3.get("final_score", 0)
            if final_r3 == 0:
                log.warning(f"    Арбитр {arb_mk} вернул final_score=0, пробуем следующего")
                continue
            log.info(f"    Арбитр {arb_mk}: {final_r3}/100 [{r3.get('quality','?')}]")
            break
        except Exception as e:
            log.warning(f"    Арбитр {arb_mk} провал: {e}")

    # ── АГРЕГАЦИЯ ─────────────────────────────────────────────────
    if r3 and "final_score" in r3 and float(r3.get("final_score", 0)) > 0:
        final_score = max(0.0, min(100.0, float(r3["final_score"])))
        criteria    = {
            "complaints":      float(r3.get("complaints",      0)),
            "disease_history": float(r3.get("disease_history", 0)),
            "comorbidities":   float(r3.get("comorbidities",   0)),
            "habits":          float(r3.get("habits",          0)),
            "labs":            float(r3.get("labs",            0)),
            "imaging":         float(r3.get("imaging",         0)),
            "penalties":       float(r3.get("penalties",       0)),
        }
        quality      = str(r3.get("quality",  "—"))
        verdict_text = str(r3.get("verdict",  ""))
    else:
        # Fallback: среднее R2
        r2_scores = [v["final_score"] for v in r2.values()
                     if v.get("final_score") is not None and v["final_score"] > 0]
        if not r2_scores:
            # Все нули — берём R1
            r2_scores = [v.get("final_score", 0) for v in valid_r1.values()
                         if v.get("final_score", 0) > 0]
        final_score = round(sum(r2_scores) / len(r2_scores), 1) if r2_scores else 0.0
        criteria    = {}
        quality     = "—"
        verdict_text = "R3 недоступен — среднее R2"
        log.warning("  R3 недоступен — использую среднее R2")

    # Флаги безопасности — если ЛЮБОЙ раунд поднял
    safety_flag = (bool(r3.get("safety_flag", False)) or
                   any(v.get("safety_flag", False) for v in {**r1, **r2}.values()))
    iodine_flag = (bool(r3.get("iodine_flag", False)) or
                   any(v.get("iodine_flag", False) for v in {**r1, **r2}.values()))
    # Quality определяется по финальному баллу
    # safety_flag лишь добавляет пометку, но не переопределяет качество
    if quality in ("—", ""):
        quality = _score_to_quality(final_score)

    all_hall = _unique([v.get("hallucinations", [])
                        for v in {**r1, **r2}.values()])

    log.info(f"\n  ✓ ИТОГ {emr_id}/{model_id}: {final_score:.1f}/100 [{quality}]"
             + (" 🚨SAFETY" if safety_flag else "")
             + (" ⚠ЙОД"    if iodine_flag  else ""))

    return {
        "emr_id":         emr_id,
        "model_id":       model_id,
        "final_score":    final_score,
        "quality":        quality,
        "safety_flag":    safety_flag,
        "iodine_flag":    iodine_flag,
        "verdict":        verdict_text,
        "criteria":       criteria,
        "hallucinations": all_hall,
        "r1_scores": {mk: v.get("final_score") for mk, v in r1.items()},
        "r2_scores": {mk: v.get("final_score") for mk, v in r2.items()},
        "r1_details": {mk: {k: v2 for k, v2 in v.items()
                            if k not in ("missing_clinical","error")}
                       for mk, v in r1.items()},
        "r2_details": r2,
    }


def _unique(lists):
    seen, out = set(), []
    for lst in lists:
        for item in lst:
            k = str(item)[:80]
            if k not in seen:
                seen.add(k)
                out.append(item)
    return out


def _err(emr_id, model_id, reason):
    return {
        "emr_id": emr_id, "model_id": model_id, "final_score": 0,
        "quality": "ошибка", "safety_flag": False, "iodine_flag": False,
        "verdict": reason, "criteria": {}, "hallucinations": [],
        "r1_scores": {}, "r2_scores": {}, "r1_details": {}, "r2_details": {},
    }


# ══════════════════════════════════════════════════════════════════
# ДАННЫЕ
# ══════════════════════════════════════════════════════════════════

def load_data(path: str) -> list[dict]:
    df = pd.read_excel(path, dtype=str).fillna("")
    required = ["emr_id", "model_id", "source_text", "summary_text"]
    missing  = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Нет колонок: {missing}. Есть: {list(df.columns)}")
    log.info(f"Загружено {len(df)} записей")
    return df.to_dict("records")


# ══════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════

QUALITY_COLORS = {
    "отличное": "C6EFCE", "хорошее": "FFEB9C",
    "удовлетворительное": "FFCC99", "неудовлетворительное": "FFC7CE",
    "опасное": "FF0000", "ошибка": "D9D9D9", "—": "FFFFFF",
}

def _brd():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _h(ws, r, c, v, bg="1F4E79", fc="FFFFFF"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=fc, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _brd()

def _d(ws, r, c, v, bold=False, bg=None, red=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, color="C00000" if red else "000000", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _brd()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)

def _sbg(s) -> Optional[str]:
    if s is None: return "D9D9D9"
    s = float(s)
    if s >= 80: return "C6EFCE"
    if s >= 60: return "FFEB9C"
    if s >= 40: return "FFCC99"
    return "FFC7CE"


def save_results(results: list[dict], path: str):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ── Лист 1: Сводная ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Сводная таблица"
    hdrs = ["ЭМК","Модель","⭐ Итог\n/100","🚨\nSafety","⚠\nЙод",
            "Жалобы\n/15","Анамнез\n/15","Сопутств.\n/20","Привычки\n/5",
            "Лаб.\n/20","Инструм.\n/25","Штрафы","Качество",
            "R1_A","R1_B","R1_C","R2_A","R2_B","R2_C",
            "Галлюцинации","Вердикт"]
    for ci, h in enumerate(hdrs, 1):
        _h(ws, 1, ci, h)
    ws.row_dimensions[1].height = 45
    ws.freeze_panes = "A2"

    for ridx, r in enumerate(results, 2):
        cr = r.get("criteria", {})
        q  = r.get("quality", "—")
        sf = r.get("safety_flag", False)
        io = r.get("iodine_flag", False)
        fs = r.get("final_score", 0)
        r1s = r.get("r1_scores", {})
        r2s = r.get("r2_scores", {})
        hall = "; ".join(str(x) for x in r.get("hallucinations",[]))[:200]

        row = [
            r["emr_id"], r["model_id"], fs,
            "🚨 ДА" if sf else "ок",
            "⚠ ДА"  if io else "нет",
            cr.get("complaints",0), cr.get("disease_history",0),
            cr.get("comorbidities",0), cr.get("habits",0),
            cr.get("labs",0), cr.get("imaging",0), cr.get("penalties",0),
            q,
            r1s.get("LLM_A"), r1s.get("LLM_B"), r1s.get("LLM_C"),
            r2s.get("LLM_A"), r2s.get("LLM_B"), r2s.get("LLM_C"),
            hall, r.get("verdict","")[:150],
        ]
        bgs = [None,None,_sbg(fs),
               "FFC7CE" if sf else None,
               "FFD966" if io else None,
               None,None,None,None,None,None,None,
               QUALITY_COLORS.get(q,"FFFFFF")] + [None]*8

        for ci, (val, bg) in enumerate(zip(row, bgs), 1):
            neg  = isinstance(val,(int,float)) and val < 0
            bold = ci == 3
            _d(ws, ridx, ci, val, bold=bold, bg=bg, red=neg)
        ws.row_dimensions[ridx].height = 18

    widths = [10,10,9,7,7,9,9,10,8,8,9,8,18,8,8,8,8,8,8,45,40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Лист 2: Детали по раундам ────────────────────────────────
    ws2 = wb.create_sheet("R1 детали")
    dh2 = ["ЭМК","Модель","Судья","Раунд","Итог","Жалобы","Анамнез",
            "Сопутств.","Привычки","Лаб.","Инструм.","Штрафы","Йод","Safety"]
    for ci, h in enumerate(dh2, 1):
        _h(ws2, 1, ci, h)
    ws2.freeze_panes = "A2"
    ridx2 = 2
    for r in results:
        for mk, det in r.get("r1_details", {}).items():
            if "error" in det:
                continue
            row2 = [
                r["emr_id"], r["model_id"], mk, "R1",
                det.get("final_score"), det.get("complaints"),
                det.get("disease_history"), det.get("comorbidities"),
                det.get("habits"), det.get("labs"), det.get("imaging"),
                det.get("penalties"),
                "⚠" if det.get("iodine_flag") else "—",
                "🚨" if det.get("safety_flag") else "—",
            ]
            for ci, v in enumerate(row2, 1):
                _d(ws2, ridx2, ci, v,
                   bg="FFC7CE" if det.get("safety_flag") and ci==14 else None,
                   red=isinstance(v,(int,float)) and v<0)
            ridx2 += 1

        for mk, det in r.get("r2_details", {}).items():
            if "error" in det:
                continue
            row2 = [
                r["emr_id"], r["model_id"], mk, "R2",
                det.get("final_score"), det.get("complaints"),
                det.get("disease_history"), det.get("comorbidities"),
                det.get("habits"), det.get("labs"), det.get("imaging"),
                det.get("penalties"),
                "⚠" if det.get("iodine_flag") else "—",
                "🚨" if det.get("safety_flag") else "—",
            ]
            for ci, v in enumerate(row2, 1):
                _d(ws2, ridx2, ci, v,
                   bg="FFEB9C" if ci == 4 else None,
                   red=isinstance(v,(int,float)) and v<0)
            ridx2 += 1

    for i, w in enumerate([8,8,7,5,7,7,7,9,7,7,8,7,5,5],1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Лист 3: Статистика ───────────────────────────────────────
    ws3 = wb.create_sheet("Статистика")
    sh = ["Модель","N","Среднее","Медиана","Мин","Макс",
          "🚨 Safety","⚠ Йод","Отл.","Хор.","Удовл.","Неудовл.","Опасных"]
    for ci, h in enumerate(sh, 1):
        _h(ws3, 1, ci, h)
    df_r = pd.DataFrame([{
        "model_id": r["model_id"], "score": r["final_score"],
        "quality": r["quality"], "safety": r.get("safety_flag",False),
        "iodine": r.get("iodine_flag",False),
    } for r in results])
    if not df_r.empty:
        for i, (mid, g) in enumerate(df_r.groupby("model_id"), 2):
            qc  = g["quality"].value_counts()
            row = [mid, len(g),
                   round(g["score"].mean(),1), round(g["score"].median(),1),
                   round(g["score"].min(),1),  round(g["score"].max(),1),
                   int(g["safety"].sum()), int(g["iodine"].sum()),
                   qc.get("отличное",0), qc.get("хорошее",0),
                   qc.get("удовлетворительное",0),
                   qc.get("неудовлетворительное",0), qc.get("опасное",0)]
            for ci, v in enumerate(row, 1):
                bg = (_sbg(v) if ci==3 else
                      ("FFC7CE" if ci in (7,8) and v>0 else None))
                _d(ws3, i, ci, v, bg=bg)
    for i, w in enumerate([16,5,9,9,8,8,8,8,8,8,8,8,8],1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Лист 4: Критические ошибки ───────────────────────────────
    ws4 = wb.create_sheet("🚨 Ошибки")
    for ci, h in enumerate(["ЭМК","Модель","Тип","Описание","Балл"],1):
        _h(ws4, 1, ci, h, bg="C00000")
    ridx4 = 2
    for r in results:
        if r.get("safety_flag"):
            for ci, v in enumerate([r["emr_id"],r["model_id"],"🚨 Safety",
                                     r.get("verdict","")[:200],r["final_score"]],1):
                _d(ws4, ridx4, ci, v, bg="FFC7CE", bold=True)
            ridx4 += 1
        if r.get("iodine_flag"):
            for ci, v in enumerate([r["emr_id"],r["model_id"],"⚠ Аллергия на йод",
                                     "Пропущена — критично для КТ с контрастом!",
                                     r["final_score"]],1):
                _d(ws4, ridx4, ci, v, bg="FFD966", bold=True)
            ridx4 += 1
        for h in r.get("hallucinations",[]):
            for ci, v in enumerate([r["emr_id"],r["model_id"],"Галлюцинация",
                                     str(h)[:200],"—"],1):
                _d(ws4, ridx4, ci, v, bg="FFEB9C")
            ridx4 += 1
    for i, w in enumerate([10,12,20,70,8],1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    log.info(f"✅ Сохранено → {path}")


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════
# ЧЕКПОИНТЫ — возможность продолжить с места остановки
# ══════════════════════════════════════════════════════════════════

CHECKPOINT_FILE = "results/checkpoint.json"
CHECKPOINT_EVERY = 5   # сохранять каждые N суммаризаций


def save_checkpoint(results: list[dict], processed_keys: set, idx: int):
    """Сохраняет прогресс в JSON-файл чекпоинта."""
    Path(CHECKPOINT_FILE).parent.mkdir(parents=True, exist_ok=True)
    data = {
        "last_idx":       idx,
        "processed_keys": list(processed_keys),
        "results":        results,
    }
    tmp = CHECKPOINT_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    Path(tmp).replace(CHECKPOINT_FILE)   # атомарная запись
    log.info(f"  💾 Чекпоинт сохранён: {idx} записей → {CHECKPOINT_FILE}")


def load_checkpoint() -> tuple[list[dict], set, int]:
    """
    Загружает чекпоинт если он существует.
    Возвращает (results, processed_keys, last_idx).
    """
    if not Path(CHECKPOINT_FILE).exists():
        return [], set(), 0
    try:
        with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        results        = data.get("results", [])
        processed_keys = set(data.get("processed_keys", []))
        last_idx       = data.get("last_idx", 0)
        log.info(f"  ♻️  Чекпоинт найден: {last_idx} записей уже обработано")
        log.info(f"      Продолжаем с записи {last_idx + 1}")
        return results, processed_keys, last_idx
    except Exception as e:
        log.warning(f"  ⚠ Ошибка чтения чекпоинта: {e} — начинаем заново")
        return [], set(), 0


def make_key(emr_id: str, model_id: str) -> str:
    """Уникальный ключ для пары ЭМК+модель."""
    return f"{emr_id}::{model_id}"


def main():
    log.info("═"*60)
    log.info("  АРЕНА LLM v3.1 — R1→R2→R3")
    log.info("═"*60)

    try:
        r = requests.get("http://localhost:11434/api/tags", timeout=5)
        available = [m["name"] for m in r.json().get("models", [])]
        log.info(f"Ollama: {len(available)} моделей")
        for mk, cfg in MODELS.items():
            name = cfg["name"]
            ok   = any(name in m for m in available)
            log.info(f"  {mk} ({name}): {'✅' if ok else '❌  → ollama pull ' + name}")
    except Exception as e:
        log.error(f"Ollama недоступен: {e}")
        log.error("Запусти приложение Ollama или: ollama serve")
        return

    try:
        records = load_data(INPUT_XLSX)
    except Exception as e:
        log.error(str(e))
        return

    # ── Загрузка чекпоинта (resume) ─────────────────────────────
    results, processed_keys, _ = load_checkpoint()

    skipped = 0
    total   = len(records)

    for idx, rec in enumerate(records, 1):
        key = make_key(rec["emr_id"], rec["model_id"])

        # Пропускаем уже обработанные записи
        if key in processed_keys:
            skipped += 1
            if skipped == 1:
                log.info(f"  ♻️  Пропускаем уже обработанные записи...")
            continue

        log.info(f"\n[{idx}/{total}] ЭМК={rec['emr_id']} Модель={rec['model_id']}")
        try:
            result = evaluate_one(
                source   = rec["source_text"],
                summary  = rec["summary_text"],
                emr_id   = rec["emr_id"],
                model_id = rec["model_id"],
            )
            results.append(result)
        except Exception as e:
            log.error(f"  ❌ {e}")
            results.append(_err(rec["emr_id"], rec["model_id"], str(e)))

        processed_keys.add(key)

        # Чекпоинт каждые N записей
        if len(processed_keys) % CHECKPOINT_EVERY == 0:
            save_checkpoint(results, processed_keys, len(processed_keys))

        # Excel каждые 10 записей
        done = len(processed_keys)
        if done % 10 == 0 or idx == total:
            save_results(results, OUTPUT_XLSX)
            log.info(f"  📊 {done}/{total} в Excel")

    # Финальное сохранение
    save_checkpoint(results, processed_keys, len(processed_keys))
    save_results(results, OUTPUT_XLSX)

    safety_n = sum(1 for r in results if r.get("safety_flag"))
    iodine_n = sum(1 for r in results if r.get("iodine_flag"))
    log.info(f"\n🏁 Готово! {len(results)}/{total} (пропущено: {skipped})")
    log.info(f"   🚨 Safety: {safety_n}  ⚠ Йод: {iodine_n}")
    log.info(f"   📊 {OUTPUT_XLSX}")
    log.info(f"   Чтобы начать заново: удали {CHECKPOINT_FILE}")


if __name__ == "__main__":
    main()
