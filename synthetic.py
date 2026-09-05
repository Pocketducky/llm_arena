"""
Блок 6 (переработан) — валидационный контур на синтетическом stress-наборе
управляемых искажений (Сеченовский ун-т).

Источник: materials/Синтетический_НД_для_тестирования_Сеченовский_университет.xlsx

СТРУКТУРА НАБОРА (установлена инспекцией; см. план / docstring ROW_DISTORTIONS):
  • лист "Лист1"; колонка "Unnamed: 0" — метки строк; 10 колонок-пациентов
    (Ж1..Ж5, М1..М5);
  • строка 0  — исходный нарратив ЭМК (~1800-2900 симв.);
  • строка 1  — ЭТАЛОННАЯ корректная суммаризация (gold, ~500 симв.);
  • строки 2..38 — та же суммаризация с ВНЕДРЁННЫМИ искажениями, организованными
    как кумулятивный градиент серьёзности по тематическим блокам (×1/×2/×3).

КЛЮЧЕВОЕ СВОЙСТВО, на котором держится вся валидация: тип искажения в строке N
ОДИНАКОВ для всех 10 пациентов (слова разные — тип один). Значит ground truth
для типа/критичности — это НОМЕР СТРОКИ, а не результат diff-инференса. Легенда
ROW_DISTORTIONS — единый редактируемый источник правды; научная команда может
переразметить пограничные строки одной правкой данных.

ДВА РЕЖИМА ОЦЕНКИ:
  A) Детекция искажений (основной критерий точности): сравниваем ЭТАЛОН (строка 1)
     с каждой искажённой строкой N и проверяем, поднял ли алгоритм флаг. По легенде:
     critical-строки ОБЯЗАНЫ флагаться (recall), benign — НЕ должны (FPR).
     Это изолирует ровно внедрённое искажение (эталон 1 -> N), в отличие от сырого
     сопоставления соседних строк (i, i+1), которое на кумулятивном наборе мерит
     «откат предыдущих правок + новая» — бессмысленный сигнал.
  B) Сквозной прод-прогон (проверка реалистичности): гоним пару ЭТАЛОН(1)->N
     через реальный конвейер judge.evaluate_summary и смотрим вердикт accept/reject.
     Референс — именно эталонная суммаризация (стр.1), а не исходная ЭМК (стр.0):
     искажения внедрены ОТНОСИТЕЛЬНО эталона, и сверка двух суммаризаций
     сопоставимого объёма даёт осмысленный шлюз. Сверка сжатой суммаризации с
     полным исходником при scope=None отклоняла даже эталон (неполное покрытие
     по конструкции — суммаризация законно короче) — вырожденный сигнал.

ДВИЖОК ДЕТЕКЦИИ — rule-based (числа/единицы/полярность) + LLM-семантика (подмена/
пропажа сущностей через JudgePanel). Rule-based ловит числа/единицы/отрицание и
расширенный набор антонимических инверсий (objective_layer._ANTONYM_PAIRS); подмену
клинических сущностей diff не различает в принципе — это работа LLM-слоя.
"""

from __future__ import annotations

import argparse
import glob
import logging
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional

import pandas as pd

import objective_layer
import reference_metrics
from llm_client import JudgePanel

log = logging.getLogger("synthetic")

SHEET_NAME = "Лист1"
ROW_LABEL_COL = "Unnamed: 0"
SOURCE_ROW = 0            # исходный нарратив ЭМК
GOLD_ROW = 1             # эталонная корректная суммаризация
FIRST_DISTORTION_ROW = 2
LAST_DISTORTION_ROW = 38

CRITICAL = "critical"
BENIGN = "benign"


# ══════════════════════════════════════════════════════════════════
# ЛЕГЕНДА ИСКАЖЕНИЙ — ЕДИНЫЙ РЕДАКТИРУЕМЫЙ ИСТОЧНИК ПРАВДЫ
# ══════════════════════════════════════════════════════════════════

# Ожидаемая ИТОГОВАЯ КАТЕГОРИЯ по типу искажения. В v1 такого отображения не
# было вовсе: легенда хранила только critical/benign, а режим B схлопывал три
# категории в бинарь `rejected = (category == "Неприемлемо")`. Из-за этого
# критическое искажение, получившее «Требует редактирования», считалось промахом
# наравне с «Готово» — понять, промах это или почти-попадание, было нельзя, а
# точность по benign не мерилась вовсе (для них проверялось только отсутствие
# отказа, а не то, признаны ли они пригодными).
CATEGORY_READY = "Готово к клиническому применению"
CATEGORY_EDIT = "Требует редактирования"
CATEGORY_REJECT = "Неприемлемо"


@dataclass(frozen=True)
class DistortionSpec:
    """Что внедрено в строке N (одинаково для всех 10 колонок-пациентов)."""
    dtype: str          # тип искажения (группирует ×1/×2/×3 строки одного блока)
    severity: str       # CRITICAL | BENIGN
    taxonomy: str       # блок таксономии дизайна (A-E) или "-"
    note: str           # пример конкретной правки (для отчёта/ревью)
    # Ожидаемый исход. По умолчанию выводится из severity: критическое искажение
    # обязано давать «Неприемлемо», безобидное — «Готово к клиническому
    # применению». Переопределяется точечно там, где научная команда решила
    # иначе (см. пограничные строки ниже).
    expected_category: Optional[str] = None
    # Ожидается ли срабатывание стоп-правила E1 (клиническая опасность).
    # Не всякое критическое искажение опасно: пропуск контента — это провал
    # полноты (блок B), а не опасное утверждение.
    expected_e1: Optional[bool] = None

    @property
    def is_critical(self) -> bool:
        return self.severity == CRITICAL

    @property
    def category_expected(self) -> str:
        if self.expected_category:
            return self.expected_category
        return CATEGORY_REJECT if self.is_critical else CATEGORY_READY

    @property
    def e1_expected(self) -> bool:
        if self.expected_e1 is not None:
            return self.expected_e1
        # По умолчанию E1 ждём от искажений, создающих ОПАСНОЕ УТВЕРЖДЕНИЕ
        # (инверсия смысла, отрицание, подмена сущности, искажение числа,
        # выдуманный факт), но не от пропусков и не от benign.
        return self.dtype in _E1_EXPECTED_TYPES


# Типы, для которых искажение порождает клинически ОПАСНОЕ утверждение —
# именно на них предохранитель E1 обязан срабатывать.
_E1_EXPECTED_TYPES = frozenset({
    "антонимическая_инверсия", "вставка_отрицания", "подмена_сущности",
    "искажение_числа", "ложная_интерпретация", "галлюцинация",
})


# Строки сгруппированы по тематическим блокам нарастающей интенсивности (×1/×2/×3).
# severity для пограничных строк (21-23, 27, 29-31) — ДЕФОЛТ, помечен в note;
# переразметка = правка одной строки здесь, остальной код не трогается.
ROW_DISTORTIONS: dict[int, DistortionSpec] = {
    # --- КРИТИЧЕСКИЕ ИСКАЖЕНИЯ (алгоритм ОБЯЗАН ловить) ---
    2:  DistortionSpec("антонимическая_инверсия", CRITICAL, "A", "включают→исключают / госпитализирован→выписан (×1)"),
    3:  DistortionSpec("антонимическая_инверсия", CRITICAL, "A", "+ купирован→спровоцирован (×2)"),
    4:  DistortionSpec("антонимическая_инверсия", CRITICAL, "A", "+ отрицательный→положительный / усиление→ослабление (×3)"),
    5:  DistortionSpec("вставка_отрицания", CRITICAL, "A", "вставлено «не» (×1)"),
    6:  DistortionSpec("вставка_отрицания", CRITICAL, "A", "вставлено «не» (×2)"),
    7:  DistortionSpec("вставка_отрицания", CRITICAL, "A", "вставлено «не» (×3)"),
    8:  DistortionSpec("подмена_сущности", CRITICAL, "A", "кардиореанимация→нейрореанимация / слабость→головокружение (×1)"),
    9:  DistortionSpec("подмена_сущности", CRITICAL, "A", "+ подмена единицы: минут→часов, лет→месяцев (×2)"),
    10: DistortionSpec("подмена_сущности", CRITICAL, "A", "+ сигарет→пачек / эпизодический→постоянный (×3)"),
    11: DistortionSpec("искажение_числа", CRITICAL, "A", "56→46 / 40→4 (×1)"),
    12: DistortionSpec("искажение_числа", CRITICAL, "A", "+ 150/90→250/90 / 2-го→1-го (×2)"),
    13: DistortionSpec("искажение_числа", CRITICAL, "A", "+ 40→5 / 86→56 (×3)"),
    # Ложная причинность искажает смысл, но не создаёт прямой угрозы пациенту:
    # ожидаем отказ по качеству, без срабатывания предохранителя E1.
    14: DistortionSpec("ложная_причинность", CRITICAL, "A", "«и курение»→«вследствие курения»",
                       expected_e1=False),
    15: DistortionSpec("ложная_интерпретация", CRITICAL, "A", "добавлено «повышена/снижено» к норм. значению (×1)"),
    16: DistortionSpec("ложная_интерпретация", CRITICAL, "A", "+ ещё одна выдуманная интерпретация (×2)"),
    17: DistortionSpec("ложная_интерпретация", CRITICAL, "A", "+ ещё одна выдуманная интерпретация (×3)"),
    18: DistortionSpec("галлюцинация", CRITICAL, "A", "добавлен выдуманный факт (отягощённая наследственность / кашель)"),
    19: DistortionSpec("галлюцинация", CRITICAL, "A", "добавлено выдуманное предложение-рекомендация"),
    # пропуск контента — ДЕФОЛТ critical (полнота, блок B); переразметьте при необходимости
    # Пропуск контента — провал ПОЛНОТЫ (блок B), а не опасное утверждение:
    # ждём отказ, но не E1. Ошибочно требовать здесь предохранитель — значит
    # заставить его срабатывать на любой сжатой суммаризации.
    21: DistortionSpec("пропуск_контента", CRITICAL, "B", "[ДЕФОЛТ critical] удалено слово/контент (×1)",
                       expected_e1=False),
    22: DistortionSpec("пропуск_контента", CRITICAL, "B", "[ДЕФОЛТ critical] удалён контент (×2)",
                       expected_e1=False),
    23: DistortionSpec("пропуск_контента", CRITICAL, "B", "[ДЕФОЛТ critical] удалён контент (×3)",
                       expected_e1=False),

    # --- БЕЗОБИДНЫЕ ИСКАЖЕНИЯ (алгоритм НЕ должен флагать) ---
    20: DistortionSpec("перестановка", BENIGN, "D", "предложение переставлено местами"),
    24: DistortionSpec("опечатка", BENIGN, "D", "опечатка в одном слове (×1)"),
    25: DistortionSpec("опечатка", BENIGN, "D", "опечатка (×2)"),
    26: DistortionSpec("опечатка", BENIGN, "D", "опечатка (×3)"),
    # нерелевантная вставка — ДЕФОЛТ benign (безвредный шум, блок C)
    # Посторонний текст безвреден клинически, но это дефект релевантности
    # (блок C): «Готово к клиническому применению» ему не положено.
    27: DistortionSpec("нерелевантная_вставка", BENIGN, "C", "[ДЕФОЛТ benign] вставлен посторонний текст («Сорта кофе…»)",
                       expected_category=CATEGORY_EDIT),
    28: DistortionSpec("пояснение", BENIGN, "D", "расшифровка аббревиатуры: КФК (креатинфосфокиназа)"),
    # интерпретация «в норме» — ДЕФОЛТ benign (как правило корректная: 126 ед/л в норме, Hb 115 снижен)
    29: DistortionSpec("интерпретация_норма", BENIGN, "-", "[ДЕФОЛТ benign] добавлено «в норме/повышено» (×1)"),
    30: DistortionSpec("интерпретация_норма", BENIGN, "-", "[ДЕФОЛТ benign] (×2)"),
    31: DistortionSpec("интерпретация_норма", BENIGN, "-", "[ДЕФОЛТ benign] (×3)"),
    # Дублирование — дефект читаемости (блок D), не ошибка факта.
    32: DistortionSpec("дублирование", BENIGN, "D", "предложение продублировано",
                       expected_category=CATEGORY_EDIT),
    33: DistortionSpec("пунктуация", BENIGN, "D", "лишняя запятая (×1)"),
    34: DistortionSpec("пунктуация", BENIGN, "D", "вопросительный знак (×2)"),
    35: DistortionSpec("пунктуация", BENIGN, "D", "восклицательный знак (×3)"),
    36: DistortionSpec("синоним", BENIGN, "-", "включают→охватывают / является→служит (×1)"),
    37: DistortionSpec("синоним", BENIGN, "-", "+ поступление→госпитализация (×2)"),
    38: DistortionSpec("синоним", BENIGN, "-", "+ купирован→остановлен (×3)"),
}

# Эталонная строка как контрольный случай: искажений нет, ожидается «Готово»
# без срабатывания E1.
GOLD_SPEC = DistortionSpec("эталон (контроль)", BENIGN, "-",
                           "неискажённая эталонная суммаризация",
                           expected_category=CATEGORY_READY, expected_e1=False)

CRITICAL_TYPES = frozenset(s.dtype for s in ROW_DISTORTIONS.values() if s.is_critical)
BENIGN_TYPES = frozenset(s.dtype for s in ROW_DISTORTIONS.values() if not s.is_critical)


# ══════════════════════════════════════════════════════════════════
# ЗАГРУЗКА НАБОРА
# ══════════════════════════════════════════════════════════════════

@dataclass
class CaseColumn:
    """Один клинический случай (колонка): исходник, эталон и искажённые варианты."""
    column: str
    source: str                       # строка 0
    gold: str                         # строка 1
    candidates: dict[int, str]        # {N: текст} для N=2..38


def _find_xlsx() -> str:
    candidates = (glob.glob("materials/*.xlsx")
                  + glob.glob("../materials/*.xlsx")
                  + glob.glob("**/Синтетич*.xlsx", recursive=True))
    candidates = [c for c in candidates if "~$" not in c]
    if not candidates:
        raise FileNotFoundError(
            "Не найден синтетический датасет (materials/*.xlsx). "
            "Передайте путь явным аргументом xlsx_path.")
    return candidates[0]


def _assert_row_labels(df, path) -> None:
    """Строгая проверка соответствия «номер строки xlsx <-> легенда».

    Весь ground truth привязан к НОМЕРУ СТРОКИ: ROW_DISTORTIONS[N] сопоставляется
    с df.iloc[N]. В самом файле никаких типов и критичности нет — только целые
    1..38 в колонке меток. Значит вставка или удаление одной строки молча
    сдвигает разметку ВСЕГО набора, и прогон продолжит считать метрики по
    неверным меткам. Единственный существовавший guard
    (check_legend_consistency) намеренно не фатален и сверяет лишь грубый класс
    диффа, поэтому сдвиг внутри одного класса он не заметит.

    Здесь падаем сразу и с объяснением: лучше остановить прогон, чем выпустить
    отчёт с перепутанными типами искажений.
    """
    labels = df[ROW_LABEL_COL]
    problems = []
    for n in range(GOLD_ROW, LAST_DISTORTION_ROW + 1):
        if n >= len(labels):
            problems.append(f"строка {n} отсутствует в файле")
            break
        raw = labels.iloc[n]
        try:
            got = int(float(raw))
        except (TypeError, ValueError):
            problems.append(f"iloc[{n}]: метка {raw!r} не число (ожидалось {n})")
            continue
        if got != n:
            problems.append(f"iloc[{n}]: метка {got} (ожидалось {n})")
    if problems:
        raise SystemExit(
            f"Разметка набора {path} не совпадает с легендой ROW_DISTORTIONS:\n  "
            + "\n  ".join(problems[:10])
            + "\n\nВесь ground truth привязан к номеру строки. Похоже, в файл "
              "добавили или удалили строку — метрики по нему считать нельзя. "
              "Либо восстановите порядок строк, либо перенумеруйте ROW_DISTORTIONS."
        )


def load_dataset(xlsx_path: Optional[str] = None, *, sheet: str = SHEET_NAME) -> list[CaseColumn]:
    """Парсит матрицу набора в список CaseColumn (по одному на пациента)."""
    path = xlsx_path or _find_xlsx()
    df = pd.read_excel(path, sheet_name=sheet)
    _assert_row_labels(df, path)
    case_columns = [c for c in df.columns if c != ROW_LABEL_COL]

    cases: list[CaseColumn] = []
    for col in case_columns:
        texts = df[col]
        source = texts.iloc[SOURCE_ROW]
        gold = texts.iloc[GOLD_ROW]
        if not isinstance(source, str) or not isinstance(gold, str):
            log.warning("synthetic: колонка %s — нет исходника/эталона, пропуск", col)
            continue
        candidates: dict[int, str] = {}
        for n in range(FIRST_DISTORTION_ROW, LAST_DISTORTION_ROW + 1):
            if n >= len(texts):
                break
            val = texts.iloc[n]
            if isinstance(val, str) and val.strip():
                candidates[n] = val
        cases.append(CaseColumn(column=str(col), source=source.strip(),
                                gold=gold.strip(), candidates=candidates))
    log.info("synthetic: загружено %d клинических случаев из %s", len(cases), path)
    return cases


@dataclass
class EvalPair:
    """Пара «референс -> искажённый кандидат» с известной (по легенде) правдой."""
    pair_id: str
    column: str
    row: int
    spec: DistortionSpec
    reference: str       # эталонная суммаризация стр.1 (оба режима A и B)
    candidate: str


def build_pairs(cases: list[CaseColumn], *, reference: str = "gold",
                with_gold_control: bool = False) -> list[EvalPair]:
    """
    Строит пары для оценки. reference="gold" — эталонная суммаризация (стр.1):
    используется И режимом A (детекция), И режимом B (прод-прогон), т.к.
    искажения внедрены относительно эталона. reference="source" (исходник стр.0)
    оставлен для будущих наборов без эталона. Только строки с известной легендой.

    with_gold_control добавляет контрольную пару «эталон против самого себя».
    Это нижняя граница здравомыслия системы: если НЕИСКАЖЁННЫЙ текст не признан
    пригодным, все остальные цифры бессмысленны. На пилоте эталон получал
    «Требует редактирования» (полный пайплайн) и «Неприемлемо» с E1 (режим
    только-судьи), и заметить это было негде — строка 1 в пары режима B
    вообще не входила.

    ВКЛЮЧАТЬ ТОЛЬКО ДЛЯ РЕЖИМА B. В режиме A (детекция искажений) сравнение
    эталона с самим собой — вырожденная пара: она по построению никогда не даст
    флаг и просто раздует знаменатель benign, бесплатно улучшив FPR. Метрику
    детекции это портит, а смысла не добавляет: проверять «не нашёл ли diff
    различий между идентичными текстами» нечего.
    """
    assert reference in ("gold", "source")
    pairs: list[EvalPair] = []
    for case in cases:
        ref = case.gold if reference == "gold" else case.source
        if with_gold_control and reference == "gold":
            pairs.append(EvalPair(
                pair_id=f"{case.column}#row{GOLD_ROW}", column=case.column, row=GOLD_ROW,
                spec=GOLD_SPEC, reference=ref, candidate=case.gold))
        for n, text in case.candidates.items():
            spec = ROW_DISTORTIONS.get(n)
            if spec is None:
                continue
            pairs.append(EvalPair(
                pair_id=f"{case.column}#row{n}", column=case.column, row=n,
                spec=spec, reference=ref, candidate=text))
    return pairs


# ══════════════════════════════════════════════════════════════════
# DIFF-КЛАССИФИКАТОР — ТОЛЬКО КАК КРОСС-ЧЕК ЛЕГЕНДЫ (НЕ ground truth)
# ══════════════════════════════════════════════════════════════════
# Грубая классификация типа правки по word-level диффу. Используется ИСКЛЮЧИТЕЛЬНО
# для guard-проверки консистентности набора (предупредить, если файл поменяли и
# строки «поехали»). Истина о типе/критичности — ROW_DISTORTIONS, а не этот код.

_NEGATION_PARTICLES = {"не", "нет", "ни", "без", "никогда", "никакой", "никаких"}
_DIGIT_RE = re.compile(r"\d")
_PUNCT_ONLY_RE = re.compile(r"^[\s.,;:!?()«»\"'\-–—]*$")
_STRIP_CHARS = ".,;:!?()«»\"'-–— "


def _norm(word: str) -> str:
    return word.strip(_STRIP_CHARS).lower()


def _is_typo(word_a: str, word_b: str) -> bool:
    wa, wb = _norm(word_a), _norm(word_b)
    if wa == wb:
        return False
    sm = SequenceMatcher(None, wa, wb, autojunk=False)
    diffs = [op for op in sm.get_opcodes() if op[0] != "equal"]
    if len(diffs) != 1:
        return False
    _, i1, i2, j1, j2 = diffs[0]
    span = max(i2 - i1, j2 - j1)
    return span <= 3 and min(len(wa), len(wb)) >= 6


def _classify_chunk(words_a: list[str], words_b: list[str]) -> str:
    joined_a, joined_b = " ".join(words_a), " ".join(words_b)
    set_a = {_norm(w) for w in words_a}
    set_b = {_norm(w) for w in words_b}
    if _DIGIT_RE.search(joined_a) or _DIGIT_RE.search(joined_b):
        return "число"
    if (set_a & _NEGATION_PARTICLES) != (set_b & _NEGATION_PARTICLES):
        return "отрицание"
    if not words_a or not words_b:
        longer = words_b if not words_a else words_a
        if len(longer) >= 3:
            return "вставка"
        if all(_PUNCT_ONLY_RE.match(w) for w in longer):
            return "перестановка"
        return "лексическая_замена"
    if len(words_a) == 1 and len(words_b) == 1 and _is_typo(words_a[0], words_b[0]):
        return "опечатка"
    if set_a == set_b and set_a:
        return "перестановка"
    if _norm(joined_a) == _norm(joined_b):
        return "перестановка"
    return "лексическая_замена"


def classify_distortion_coarse(text_a: str, text_b: str) -> tuple[str, ...]:
    """Грубые типы правок между text_a и text_b (для guard-кросс-чека)."""
    words_a, words_b = text_a.split(), text_b.split()
    norm_a, norm_b = [_norm(w) for w in words_a], [_norm(w) for w in words_b]
    matcher = SequenceMatcher(None, norm_a, norm_b, autojunk=False)
    raw_chunks: list[tuple[list[str], list[str]]] = []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            continue
        raw_chunks.append((words_a[i1:i2], words_b[j1:j2]))
    if not raw_chunks:
        return ("перестановка",)
    removed = Counter(_norm(w) for ca, _ in raw_chunks for w in ca if _norm(w))
    added = Counter(_norm(w) for _, cb in raw_chunks for w in cb if _norm(w))
    if removed and removed == added:
        return ("перестановка",)
    found = [_classify_chunk(ca, cb) for ca, cb in raw_chunks]
    return tuple(sorted(set(found)))


def check_legend_consistency(cases: list[CaseColumn]) -> list[str]:
    """
    Guard: убеждаемся, что набор соответствует легенде по СТРУКТУРЕ
    (10 случаев, строки на месте, кандидаты непусты) и грубо — по типу
    (diff(gold->candidate) не противоречит критичности легенды). Возвращает
    список предупреждений (пустой = всё консистентно). НЕ падает — набор может
    законно содержать пограничные случаи; задача guard — поднять флаг при ДРЕЙФЕ.
    """
    warnings: list[str] = []
    if len(cases) != 10:
        warnings.append(f"ожидалось 10 клинических случаев, получено {len(cases)}")

    # Для каждой строки N: тип правки должен быть СОГЛАСОВАН по всем колонкам
    # (грубо). Если в одной строке у разных пациентов радикально разные грубые
    # типы — возможно, файл «поехал».
    for n in range(FIRST_DISTORTION_ROW, LAST_DISTORTION_ROW + 1):
        spec = ROW_DISTORTIONS.get(n)
        if spec is None:
            continue
        coarse_per_col: list[tuple[str, ...]] = []
        for case in cases:
            cand = case.candidates.get(n)
            if cand is None:
                continue
            coarse_per_col.append(classify_distortion_coarse(case.gold, cand))
        if not coarse_per_col:
            warnings.append(f"строка {n} ({spec.dtype}): нет кандидатов ни в одной колонке")
            continue
        # benign-строки не должны массово выглядеть как «число/отрицание/вставка»
        coarse_critical_markers = {"число", "отрицание", "вставка"}
        crit_hits = sum(1 for c in coarse_per_col
                        if coarse_critical_markers & set(c))
        if not spec.is_critical and crit_hits > len(coarse_per_col) // 2:
            warnings.append(
                f"строка {n} помечена BENIGN ({spec.dtype}), но diff у "
                f"{crit_hits}/{len(coarse_per_col)} колонок выглядит критично — проверьте легенду")
    return warnings


# ══════════════════════════════════════════════════════════════════
# РЕЖИМ A — ДЕТЕКЦИЯ ИСКАЖЕНИЙ
# ══════════════════════════════════════════════════════════════════

@dataclass
class DetectionResult:
    pair: EvalPair
    report: "objective_layer.ObjectiveComparisonReport"
    caught_numeric: bool
    caught_unit: bool
    caught_polarity: bool
    caught_entity: bool
    caught_causality: bool
    caught_interpretation: bool

    @property
    def flagged(self) -> bool:
        return (self.caught_numeric or self.caught_unit or self.caught_polarity
                or self.caught_entity or self.caught_causality
                or self.caught_interpretation)

    def signals(self) -> list[str]:
        out = []
        if self.caught_numeric:
            out.append("числа")
        if self.caught_unit:
            out.append("единицы")
        if self.caught_polarity:
            out.append("полярность")
        if self.caught_entity:
            out.append("сущности")
        if self.caught_causality:
            out.append("причинность")
        if self.caught_interpretation:
            out.append("интерпретация")
        return out


# Слова, которые сами по себе НЕ идентифицируют клиническую сущность —
# фантомные missing/extra только из них не считаем содержательным сигналом.
_GENERIC_ENTITY_WORDS = frozenset({
    "анализ", "крови", "кровь", "исследование", "исследования", "данные",
    "пациент", "пациента", "пациентка", "пациентки", "больной", "больная",
    "осмотр", "состояние", "норма", "результат", "результаты",
})


def _significant_entity(name: str) -> bool:
    """Содержательна ли сущность как сигнал расхождения: не служебно-короткая
    и не состоит целиком из общих немедицинских слов. Отсекает мусорные
    извлечения LLM, которые иначе раздували бы ложные срабатывания."""
    tokens = [t for t in re.findall(r"[а-яёa-z0-9]+", name.lower())
              if len(t) >= 4 and t not in _GENERIC_ENTITY_WORDS]
    return bool(tokens)


def _entity_changed(report) -> bool:
    """Сущность пропала или подменена/добавлена — сигнал режима A.
    Флаг ставится на содержательное расхождение (после устойчивого нечёткого
    сопоставления в compare_entity_sets опечатки/словоформы/расшифровки уже
    не порождают фантомных missing/extra; здесь дополнительно отбрасываем
    мусорные/служебные сущности). Ловит подмену отделения/диагноза (стр.8-10),
    галлюцинацию (лишняя сущность) и пропуск (пропавшая сущность)."""
    if not report.entities:
        return False
    for cat, rep in report.entities.items():
        if any(_significant_entity(e) for e in rep["missing_in_b"]):
            return True
        if any(_significant_entity(e) for e in rep["extra_in_b"]):
            return True
    return False


def run_detection(pairs: list[EvalPair], *, panel: Optional[JudgePanel] = None) -> list[DetectionResult]:
    """Режим A: сравнить эталон с каждым кандидатом, зафиксировать поднятые флаги."""
    results: list[DetectionResult] = []
    for i, p in enumerate(pairs, 1):
        rep = objective_layer.compare_texts(p.reference, p.candidate, panel=panel)
        results.append(DetectionResult(
            pair=p, report=rep,
            caught_numeric=rep.numeric.get("mismatch_count", 0) > 0,
            caught_unit=rep.numeric.get("unit_mismatch_count", 0) > 0,
            caught_polarity=rep.polarity.get("flip_count", 0) > 0,
            caught_entity=_entity_changed(rep),
            caught_causality=bool(rep.causality and rep.causality.get("introduced", 0) > 0),
            caught_interpretation=bool(rep.interpretation and rep.interpretation.get("count", 0) > 0),
        ))
        if i % 50 == 0:
            log.info("synthetic[A]: обработано %d/%d пар", i, len(pairs))
    log.info("synthetic[A]: детекция завершена — %d пар", len(results))
    return results


# ══════════════════════════════════════════════════════════════════
# РЕЖИМ B — СКВОЗНОЙ ПРОД-ПРОГОН
# ══════════════════════════════════════════════════════════════════

@dataclass
class ProductionResult:
    pair: EvalPair
    category: str
    gate_status: str
    rejected: bool
    e1: bool = False
    decision_path: list = field(default_factory=list)

    @property
    def expected_category(self) -> str:
        return self.pair.spec.category_expected

    @property
    def category_correct(self) -> bool:
        return self.category == self.expected_category

    @property
    def e1_correct(self) -> bool:
        return self.e1 == self.pair.spec.e1_expected


def run_production(pairs: list[EvalPair], *, panel: JudgePanel,
                   scope: Optional[str] = None) -> list[ProductionResult]:
    """Режим B: гоним эталон->кандидат через реальный конвейер judge.evaluate_summary
    (референс — эталонная суммаризация стр.1, см. build_pairs).

    `scope` — тот же параметр, что и в прод-прогоне. Раньше был жёстко зашит
    None, из-за чего валидация шла по ДРУГОМУ набору правил шлюза, чем прод.
    """
    import aggregator
    import judge  # тяжёлый импорт — только если режим B реально запускают
    results: list[ProductionResult] = []
    for i, p in enumerate(pairs, 1):
        log.info("synthetic[B]: %d/%d %s (%s)", i, len(pairs), p.pair_id, p.spec.dtype)
        ev = judge.evaluate_summary(p.reference, p.candidate,
                                    emr_id=p.column, model_id=f"row{p.row}",
                                    scope=scope, panel=panel)
        ev = aggregator.finalize(ev)     # категорию выносит код, а не самооценка R3
        gate_status = (ev.get("gate") or {}).get("status", "?")
        category = ev.get("category", "?")
        # Служебные статусы («ошибка», «Оценка неполна») — НЕ отказ по существу:
        # засчитывать их как reject значило бы мерить надёжность инфраструктуры
        # вместо качества суммаризации.
        rejected = category == aggregator.CATEGORY_REJECT
        results.append(ProductionResult(
            pair=p, category=category, gate_status=gate_status, rejected=rejected,
            e1=bool(ev.get("e1_triggered")),
            decision_path=list(ev.get("decision_path") or [])))
    log.info("synthetic[B]: прод-прогон завершён — %d пар", len(results))
    return results


# ══════════════════════════════════════════════════════════════════
# СВОДКА И МЕТРИКИ
# ══════════════════════════════════════════════════════════════════

@dataclass
class TypeStats:
    dtype: str
    is_critical: bool
    taxonomy: str
    n: int
    flagged: int                       # сколько раз алгоритм поднял флаг
    by_signal: dict[str, int] = field(default_factory=dict)

    @property
    def rate(self) -> float:
        return round(self.flagged / self.n, 4) if self.n else 0.0

    @property
    def rate_label(self) -> str:
        return "recall" if self.is_critical else "FPR"


@dataclass
class DetectionSummary:
    total: int
    by_type: dict[str, TypeStats]
    overall_critical_recall: Optional[float]
    overall_benign_fpr: Optional[float]
    misses: list[DetectionResult]          # critical не пойман
    false_alarms: list[DetectionResult]    # benign ложно пойман


def summarize_detection(results: list[DetectionResult]) -> DetectionSummary:
    by_type: dict[str, TypeStats] = {}
    seen_types = sorted({r.pair.spec.dtype for r in results})
    for dtype in seen_types:
        subset = [r for r in results if r.pair.spec.dtype == dtype]
        spec = subset[0].pair.spec
        by_type[dtype] = TypeStats(
            dtype=dtype, is_critical=spec.is_critical, taxonomy=spec.taxonomy,
            n=len(subset), flagged=sum(1 for r in subset if r.flagged),
            by_signal={
                "числа": sum(1 for r in subset if r.caught_numeric),
                "единицы": sum(1 for r in subset if r.caught_unit),
                "полярность": sum(1 for r in subset if r.caught_polarity),
                "сущности": sum(1 for r in subset if r.caught_entity),
                "причинность": sum(1 for r in subset if r.caught_causality),
                "интерпретация": sum(1 for r in subset if r.caught_interpretation),
            },
        )
    crit = [r for r in results if r.pair.spec.is_critical]
    ben = [r for r in results if not r.pair.spec.is_critical]
    return DetectionSummary(
        total=len(results), by_type=by_type,
        overall_critical_recall=(round(sum(1 for r in crit if r.flagged) / len(crit), 4) if crit else None),
        overall_benign_fpr=(round(sum(1 for r in ben if r.flagged) / len(ben), 4) if ben else None),
        misses=[r for r in crit if not r.flagged],
        false_alarms=[r for r in ben if r.flagged],
    )


@dataclass
class ProductionSummary:
    total: int
    by_type: dict[str, TypeStats]
    overall_critical_reject: Optional[float]
    overall_benign_reject: Optional[float]
    # Матрица ошибок 3x3 и метрики по категориям. В v1 их не было: три категории
    # схлопывались в бинарь `rejected`, поэтому критическое искажение, получившее
    # «Требует редактирования», считалось промахом наравне с «Готово», а точность
    # по benign не измерялась вовсе.
    confusion: dict = field(default_factory=dict)       # {ожидаемая: {полученная: n}}
    per_category: dict = field(default_factory=dict)    # {категория: {precision, recall, f1, support}}
    accuracy: Optional[float] = None
    e1_accuracy: Optional[float] = None
    e1_confusion: dict = field(default_factory=dict)    # TP/FP/FN/TN по стоп-правилу
    service_statuses: dict = field(default_factory=dict)  # «ошибка»/«Оценка неполна»


_CATEGORY_ORDER = (CATEGORY_READY, CATEGORY_EDIT, CATEGORY_REJECT)


def summarize_production(results: list[ProductionResult]) -> ProductionSummary:
    by_type: dict[str, TypeStats] = {}
    seen_types = sorted({r.pair.spec.dtype for r in results})
    for dtype in seen_types:
        subset = [r for r in results if r.pair.spec.dtype == dtype]
        spec = subset[0].pair.spec
        by_type[dtype] = TypeStats(
            dtype=dtype, is_critical=spec.is_critical, taxonomy=spec.taxonomy,
            n=len(subset), flagged=sum(1 for r in subset if r.rejected),
            by_signal={},
        )
    crit = [r for r in results if r.pair.spec.is_critical]
    ben = [r for r in results if not r.pair.spec.is_critical]

    # Матрица ошибок: ожидаемая категория x полученная. Служебные статусы
    # («ошибка», «Оценка неполна») выносим отдельно — это отсутствие результата,
    # а не ошибочная классификация, и смешивать их с промахами нельзя.
    service = {}
    confusion: dict = {exp: {got: 0 for got in _CATEGORY_ORDER} for exp in _CATEGORY_ORDER}
    scored = []
    for r in results:
        if r.category not in _CATEGORY_ORDER:
            service[r.category] = service.get(r.category, 0) + 1
            continue
        scored.append(r)
        confusion[r.expected_category][r.category] += 1

    per_category = {}
    for cat in _CATEGORY_ORDER:
        tp = confusion[cat][cat]
        fp = sum(confusion[e][cat] for e in _CATEGORY_ORDER if e != cat)
        fn = sum(confusion[cat][g] for g in _CATEGORY_ORDER if g != cat)
        support = tp + fn
        precision = round(tp / (tp + fp), 4) if (tp + fp) else None
        recall = round(tp / support, 4) if support else None
        f1 = (round(2 * precision * recall / (precision + recall), 4)
              if precision and recall else (0.0 if support else None))
        per_category[cat] = {"precision": precision, "recall": recall,
                             "f1": f1, "support": support}

    accuracy = (round(sum(1 for r in scored if r.category_correct) / len(scored), 4)
                if scored else None)

    # Стоп-правило E1 меряем отдельно: это предохранитель безопасности, и его
    # ошибки первого и второго рода имеют разную цену.
    e1_tp = sum(1 for r in scored if r.e1 and r.pair.spec.e1_expected)
    e1_fp = sum(1 for r in scored if r.e1 and not r.pair.spec.e1_expected)
    e1_fn = sum(1 for r in scored if not r.e1 and r.pair.spec.e1_expected)
    e1_tn = sum(1 for r in scored if not r.e1 and not r.pair.spec.e1_expected)

    return ProductionSummary(
        total=len(results), by_type=by_type,
        overall_critical_reject=(round(sum(1 for r in crit if r.rejected) / len(crit), 4) if crit else None),
        overall_benign_reject=(round(sum(1 for r in ben if r.rejected) / len(ben), 4) if ben else None),
        confusion=confusion, per_category=per_category, accuracy=accuracy,
        e1_accuracy=(round((e1_tp + e1_tn) / len(scored), 4) if scored else None),
        e1_confusion={"TP": e1_tp, "FP": e1_fp, "FN": e1_fn, "TN": e1_tn},
        service_statuses=service,
    )


# ══════════════════════════════════════════════════════════════════
# КОНСОЛЬНЫЙ ОТЧЁТ
# ══════════════════════════════════════════════════════════════════

def render_detection(summary: DetectionSummary, *, llm_used: bool) -> str:
    lines = [
        "=" * 72,
        "РЕЖИМ A — ДЕТЕКЦИЯ ИСКАЖЕНИЙ (эталон стр.1 -> искажённая строка N)",
        f"Движок: rule-based (числа/единицы/полярность)" + (" + LLM-сущности" if llm_used else " (без LLM)"),
        f"Всего пар: {summary.total}",
        "=" * 72,
        "",
        "--- КРИТИЧЕСКИЕ типы (цель: высокий RECALL — обязаны ловиться) ---",
    ]
    crit = [s for s in summary.by_type.values() if s.is_critical]
    ben = [s for s in summary.by_type.values() if not s.is_critical]
    for s in sorted(crit, key=lambda x: -x.n):
        lines.append(f"  {s.dtype:<24}[{s.taxonomy}] n={s.n:<3} recall={s.rate:>6.1%} "
                     f"({s.flagged}/{s.n})  сигналы={s.by_signal}")
    if summary.overall_critical_recall is not None:
        lines.append(f"  >>> ИТОГО recall по критическим: {summary.overall_critical_recall:.1%} "
                     f"(целевой ориентир дизайна ≥ 90%)")
    lines += ["", "--- БЕЗОБИДНЫЕ типы (цель: низкий FALSE-POSITIVE RATE) ---"]
    for s in sorted(ben, key=lambda x: -x.n):
        lines.append(f"  {s.dtype:<24}[{s.taxonomy}] n={s.n:<3} FPR={s.rate:>6.1%} "
                     f"({s.flagged}/{s.n})  сигналы={s.by_signal}")
    if summary.overall_benign_fpr is not None:
        lines.append(f"  >>> ИТОГО FPR по безобидным: {summary.overall_benign_fpr:.1%}")

    if summary.misses:
        lines += ["", f"--- ПРОМАХИ: критические искажения НЕ пойманы ({len(summary.misses)}) ---"]
        for r in summary.misses[:20]:
            lines.append(f"  [{r.pair.pair_id}] {r.pair.spec.dtype}: {r.pair.spec.note}")
        if len(summary.misses) > 20:
            lines.append(f"  … ещё {len(summary.misses) - 20}")
    if summary.false_alarms:
        lines += ["", f"--- ЛОЖНЫЕ СРАБАТЫВАНИЯ: benign ошибочно пойман ({len(summary.false_alarms)}) ---"]
        for r in summary.false_alarms[:20]:
            lines.append(f"  [{r.pair.pair_id}] {r.pair.spec.dtype} ({r.signals()}): {r.pair.spec.note}")
        if len(summary.false_alarms) > 20:
            lines.append(f"  … ещё {len(summary.false_alarms) - 20}")
    return "\n".join(lines)


def render_production(summary: ProductionSummary) -> str:
    lines = [
        "=" * 78,
        "РЕЖИМ B — СКВОЗНОЙ ПРОД-ПРОГОН (эталон стр.1 -> строка N, judge.evaluate_summary)",
        f"Всего пар: {summary.total}",
        "=" * 78,
    ]

    if summary.service_statuses:
        lines += ["", "--- ПАРЫ БЕЗ РЕЗУЛЬТАТА (не участвуют в метриках) ---"]
        for status, n in sorted(summary.service_statuses.items()):
            lines.append(f"  {status:<24} {n}")
        lines.append("  (это сбои прогона, а не ошибки классификации — их надо перепрогнать)")

    # ── Матрица ошибок ────────────────────────────────────────────
    lines += ["", "--- МАТРИЦА ОШИБОК (строки — ожидаемая категория, столбцы — полученная) ---"]
    short = {CATEGORY_READY: "Готово", CATEGORY_EDIT: "Треб.ред.", CATEGORY_REJECT: "Неприемл."}
    header = " " * 26 + "".join(f"{short[c]:>12}" for c in _CATEGORY_ORDER) + f"{'всего':>9}"
    lines.append(header)
    for exp in _CATEGORY_ORDER:
        rowvals = [summary.confusion.get(exp, {}).get(got, 0) for got in _CATEGORY_ORDER]
        mark = " <- ожидалось"
        lines.append(f"  {short[exp]:<24}" + "".join(f"{v:>12}" for v in rowvals)
                     + f"{sum(rowvals):>9}" + (mark if sum(rowvals) else ""))
    if summary.accuracy is not None:
        lines.append(f"  >>> ТОЧНОСТЬ (доля совпавших категорий): {summary.accuracy:.1%}")

    # ── Метрики по категориям ─────────────────────────────────────
    lines += ["", "--- PRECISION / RECALL / F1 ПО КАТЕГОРИЯМ ---",
              f"  {'категория':<34}{'precision':>11}{'recall':>9}{'F1':>8}{'support':>9}"]
    for cat in _CATEGORY_ORDER:
        m = summary.per_category.get(cat) or {}
        fmt = lambda v: f"{v:.2f}" if isinstance(v, (int, float)) else "—"
        lines.append(f"  {cat:<34}{fmt(m.get('precision')):>11}{fmt(m.get('recall')):>9}"
                     f"{fmt(m.get('f1')):>8}{m.get('support', 0):>9}")

    # ── Стоп-правило E1 ───────────────────────────────────────────
    e1 = summary.e1_confusion or {}
    lines += ["", "--- СТОП-ПРАВИЛО E1 (предохранитель безопасности) ---",
              f"  сработал верно (TP):        {e1.get('TP', 0)}",
              f"  ложное срабатывание (FP):   {e1.get('FP', 0)}   <- цена: отвергнутая годная сводка",
              f"  пропуск опасности (FN):     {e1.get('FN', 0)}   <- цена: пропущенная угроза пациенту",
              f"  верно молчал (TN):          {e1.get('TN', 0)}"]
    if summary.e1_accuracy is not None:
        lines.append(f"  >>> ТОЧНОСТЬ E1: {summary.e1_accuracy:.1%}")
    total_e1 = sum(e1.values()) or 1
    if (e1.get("TP", 0) + e1.get("FP", 0)) == total_e1:
        lines.append("  ⚠ E1 сработал на ВСЕХ парах — предохранитель, срабатывающий всегда, "
                     "не несёт информации (ровно этот дефект дал 38/38 на пилоте)")

    # ── Разбивка по типам (как раньше) ────────────────────────────
    crit = [st for st in summary.by_type.values() if st.is_critical]
    ben = [st for st in summary.by_type.values() if not st.is_critical]
    lines += ["", "--- КРИТИЧЕСКИЕ типы (цель: высокая доля REJECT) ---"]
    for st in sorted(crit, key=lambda x: -x.n):
        lines.append(f"  {st.dtype:<24}[{st.taxonomy}] n={st.n:<3} reject={st.rate:>6.1%} ({st.flagged}/{st.n})")
    if summary.overall_critical_reject is not None:
        lines.append(f"  >>> ИТОГО reject по критическим: {summary.overall_critical_reject:.1%}")
    lines += ["", "--- БЕЗОБИДНЫЕ типы (цель: НИЗКАЯ доля REJECT) ---"]
    for st in sorted(ben, key=lambda x: -x.n):
        lines.append(f"  {st.dtype:<24}[{st.taxonomy}] n={st.n:<3} reject={st.rate:>6.1%} ({st.flagged}/{st.n})")
    if summary.overall_benign_reject is not None:
        lines.append(f"  >>> ИТОГО reject по безобидным: {summary.overall_benign_reject:.1%}")
    lines.append("=" * 78)
    return "\n".join(lines)


def compute_reference_metrics(pairs: list[EvalPair], *, use_bertscore: bool) -> dict[str, dict]:
    """Возвращает {pair_id: {"rouge_l": F1, "bertscore": F1|None}} по парам
    эталон↔кандидат. BERTScore считается батчем, только если пакет установлен."""
    per_pair = {p.pair_id: {"rouge_l": reference_metrics.rouge_l(p.reference, p.candidate)["f1"],
                            "bertscore": None}
                for p in pairs}
    if use_bertscore:
        if reference_metrics.bertscore_available():
            log.info("synthetic[A]: BERTScore по %d парам (загрузка модели)...", len(pairs))
            bs = reference_metrics.bertscore([p.reference for p in pairs],
                                             [p.candidate for p in pairs])
            for p, b in zip(pairs, bs):
                per_pair[p.pair_id]["bertscore"] = b["f1"]
        else:
            log.warning("synthetic[A]: BERTScore пропущен — пакет bert-score не установлен "
                        "(pip install -r requirements-metrics.txt)")
    return per_pair


def summarize_reference_metrics(det_results: list[DetectionResult],
                                per_pair: dict[str, dict]) -> dict:
    """Средние ROUGE-L и BERTScore F1 по каждому типу искажения + по critical/benign."""
    def _mean(xs):
        xs = [x for x in xs if x is not None]
        return round(sum(xs) / len(xs), 4) if xs else None

    by_type: dict[str, dict] = {}
    crit_r, crit_b, ben_r, ben_b = [], [], [], []
    for r in det_results:
        m = per_pair.get(r.pair.pair_id)
        if not m:
            continue
        bucket = by_type.setdefault(r.pair.spec.dtype,
                                    {"is_critical": r.pair.spec.is_critical, "rouge": [], "bert": []})
        bucket["rouge"].append(m["rouge_l"])
        bucket["bert"].append(m["bertscore"])
        (crit_r if r.pair.spec.is_critical else ben_r).append(m["rouge_l"])
        (crit_b if r.pair.spec.is_critical else ben_b).append(m["bertscore"])
    by_type_out = {dt: {"is_critical": v["is_critical"], "n": len(v["rouge"]),
                        "rouge_l": _mean(v["rouge"]), "bertscore": _mean(v["bert"])}
                   for dt, v in by_type.items()}
    return {
        "by_type": by_type_out,
        "overall": {"critical": {"rouge_l": _mean(crit_r), "bertscore": _mean(crit_b)},
                    "benign": {"rouge_l": _mean(ben_r), "bertscore": _mean(ben_b)}},
        "bertscore_used": any(m["bertscore"] is not None for m in per_pair.values()),
    }


def render_reference_metrics(summary: dict) -> str:
    bs = summary["bertscore_used"]
    lines = [
        "=" * 72,
        "REFERENCE-МЕТРИКИ (эталон стр.1 ↔ кандидат) — ROUGE-L"
        + (" + BERTScore" if bs else " (BERTScore отключён/не установлен)"),
        "Выше = ближе к эталону; benign-правки должны быть близки к 1, искажения — ниже",
        "=" * 72,
        "",
    ]
    crit = [(dt, v) for dt, v in summary["by_type"].items() if v["is_critical"]]
    ben = [(dt, v) for dt, v in summary["by_type"].items() if not v["is_critical"]]

    def _fmt(val):
        return f"{val:.3f}" if val is not None else "  —  "

    for label, group in (("--- КРИТИЧЕСКИЕ типы ---", crit), ("--- БЕЗОБИДНЫЕ типы ---", ben)):
        lines.append(label)
        for dt, v in sorted(group, key=lambda kv: -kv[1]["n"]):
            lines.append(f"  {dt:<26} n={v['n']:<3} ROUGE-L={_fmt(v['rouge_l'])}"
                         + (f"  BERTScore={_fmt(v['bertscore'])}" if bs else ""))
        lines.append("")
    o = summary["overall"]
    lines.append(f">>> ИТОГО critical: ROUGE-L={_fmt(o['critical']['rouge_l'])}"
                 + (f"  BERTScore={_fmt(o['critical']['bertscore'])}" if bs else ""))
    lines.append(f">>> ИТОГО benign:   ROUGE-L={_fmt(o['benign']['rouge_l'])}"
                 + (f"  BERTScore={_fmt(o['benign']['bertscore'])}" if bs else ""))
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════
# EXCEL-ОТЧЁТ
# ══════════════════════════════════════════════════════════════════

def write_excel(path: str, det: Optional[DetectionSummary],
                det_results: Optional[list[DetectionResult]],
                prod: Optional[ProductionSummary],
                ref_summary: Optional[dict] = None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    bold = Font(bold=True)
    crit_fill = PatternFill("solid", fgColor="F8CBAD")
    ben_fill = PatternFill("solid", fgColor="C6E0B4")

    ws = wb.active
    ws.title = "Сводка по типам (A)"
    if det is not None:
        ws.append(["Тип искажения", "Критичность", "Блок", "N", "Поймано",
                   "Метрика", "Значение", "числа", "единицы", "полярность",
                   "сущности", "причинность", "интерпретация"])
        for c in ws[1]:
            c.font = bold
        for s in sorted(det.by_type.values(), key=lambda x: (not x.is_critical, -x.n)):
            ws.append([s.dtype, "critical" if s.is_critical else "benign", s.taxonomy,
                       s.n, s.flagged, s.rate_label, s.rate,
                       s.by_signal.get("числа", 0), s.by_signal.get("единицы", 0),
                       s.by_signal.get("полярность", 0), s.by_signal.get("сущности", 0),
                       s.by_signal.get("причинность", 0), s.by_signal.get("интерпретация", 0)])
            ws[ws.max_row][0].fill = crit_fill if s.is_critical else ben_fill
        ws.append([])
        ws.append(["ИТОГО recall (critical)", "", "", "", "", "", det.overall_critical_recall])
        ws.append(["ИТОГО FPR (benign)", "", "", "", "", "", det.overall_benign_fpr])

    if det_results is not None:
        ws2 = wb.create_sheet("Детали пар (A)")
        ws2.append(["pair_id", "колонка", "строка", "тип", "критичность",
                    "флаг", "сигналы", "верно?", "примечание"])
        for c in ws2[1]:
            c.font = bold
        for r in det_results:
            correct = (r.flagged == r.pair.spec.is_critical)
            ws2.append([r.pair.pair_id, r.pair.column, r.pair.row, r.pair.spec.dtype,
                        "critical" if r.pair.spec.is_critical else "benign",
                        "ДА" if r.flagged else "нет", ", ".join(r.signals()),
                        "OK" if correct else "ОШИБКА", r.pair.spec.note])

    if prod is not None:
        ws3 = wb.create_sheet("Сводка по типам (B)")
        ws3.append(["Тип искажения", "Критичность", "Блок", "N", "Reject", "Доля reject"])
        for c in ws3[1]:
            c.font = bold
        for s in sorted(prod.by_type.values(), key=lambda x: (not x.is_critical, -x.n)):
            ws3.append([s.dtype, "critical" if s.is_critical else "benign", s.taxonomy,
                        s.n, s.flagged, s.rate])
            ws3[ws3.max_row][0].fill = crit_fill if s.is_critical else ben_fill

    if ref_summary is not None:
        ws4 = wb.create_sheet("Reference-метрики (A)")
        ws4.append(["Тип искажения", "Критичность", "N", "ROUGE-L F1 (ср.)", "BERTScore F1 (ср.)"])
        for c in ws4[1]:
            c.font = bold
        for dt, v in sorted(ref_summary["by_type"].items(),
                            key=lambda kv: (not kv[1]["is_critical"], -kv[1]["n"])):
            ws4.append([dt, "critical" if v["is_critical"] else "benign", v["n"],
                        v["rouge_l"], v["bertscore"]])
            ws4[ws4.max_row][0].fill = crit_fill if v["is_critical"] else ben_fill
        ws4.append([])
        o = ref_summary["overall"]
        ws4.append(["ИТОГО critical", "", "", o["critical"]["rouge_l"], o["critical"]["bertscore"]])
        ws4.append(["ИТОГО benign", "", "", o["benign"]["rouge_l"], o["benign"]["bertscore"]])

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    log.info("synthetic: Excel-отчёт сохранён -> %s", path)


# ══════════════════════════════════════════════════════════════════
# CLI / ОРКЕСТРАЦИЯ
# ══════════════════════════════════════════════════════════════════

def run(*, mode: str = "a", limit_cols: Optional[int] = None,
        use_llm: bool = True, profile: Optional[str] = None,
        xlsx_path: Optional[str] = None, excel_out: Optional[str] = None,
        ref_metrics: bool = False, scope: Optional[str] = None) -> dict:
    cases = load_dataset(xlsx_path)
    warnings = check_legend_consistency(cases)
    for w in warnings:
        log.warning("guard легенды: %s", w)
    if limit_cols is not None:
        cases = cases[:limit_cols]
        log.info("synthetic: ограничение до %d колонок", len(cases))

    panel = JudgePanel(profile) if use_llm else None
    out: dict = {"warnings": warnings}

    det_summary = det_results = prod_summary = ref_summary = None
    if mode in ("a", "both"):
        pairs_a = build_pairs(cases, reference="gold")
        det_results = run_detection(pairs_a, panel=panel)
        det_summary = summarize_detection(det_results)
        print(render_detection(det_summary, llm_used=use_llm))
        out["detection"] = det_summary
        if ref_metrics:
            per_pair = compute_reference_metrics(pairs_a, use_bertscore=True)
            ref_summary = summarize_reference_metrics(det_results, per_pair)
            print("\n" + render_reference_metrics(ref_summary))
            out["reference_metrics"] = ref_summary
    if mode in ("b", "both"):
        if panel is None:
            panel = JudgePanel(profile)   # режим B без LLM невозможен
        # reference="gold" (эталонная суммаризация стр.1), а НЕ "source": искажения
        # внедрены относительно эталона, и «правильно» здесь означает «как в эталоне».
        # Сверка сжатой суммаризации (~500 симв.) с полной ЭМК-исходником (~2500 симв.)
        # при scope=None систематически проваливала шлюз (entity_recall<0.5, неполное
        # покрытие чисел/предикатов) для ЛЮБОЙ суммаризации, включая эталонную, —
        # отсюда прежний вырожденный 100% reject. Эталон-референс делает сверку
        # сопоставимой по объёму: benign-кандидат ≈ эталон проходит шлюз и судей,
        # critical расходится и отклоняется. См. docstring модуля, режим B.
        pairs_b = build_pairs(cases, reference="gold", with_gold_control=True)
        prod_results = run_production(pairs_b, panel=panel, scope=scope)
        prod_summary = summarize_production(prod_results)
        print("\n" + render_production(prod_summary))
        out["production"] = prod_summary

    if excel_out:
        write_excel(excel_out, det_summary, det_results, prod_summary, ref_summary)
        out["excel"] = excel_out
    return out


def _build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--mode", choices=["a", "b", "both"], default="a",
                   help="a — детекция (эталон->искажение, быстро); b — прод-прогон "
                        "(эталон->строка через judge, медленно); both — оба")
    p.add_argument("--limit-cols", type=int, default=None,
                   help="ограничить число клинических случаев (колонок) — для быстрой пробы")
    p.add_argument("--no-llm", action="store_true",
                   help="режим A без LLM-сущностей (только rule-based; режим B всё равно требует LLM)")
    p.add_argument("--scope", default=None,
                   help="задача/адресат суммаризации для режима B. Пусто = None "
                        "(пары «эталон vs искажение» сопоставимы по объёму, сужать нечего). "
                        "Задавайте тот же scope, что и в прод-прогоне, если хотите мерить "
                        "ровно прод-конфигурацию шлюза.")
    p.add_argument("--profile", default=None, help="профиль моделей (по умолчанию config.ACTIVE_PROFILE)")
    p.add_argument("--xlsx", default=None, help="путь к xlsx (по умолчанию — автопоиск в materials/)")
    p.add_argument("--excel-out", default=None, help="куда сохранить Excel-отчёт (например reports/synthetic.xlsx)")
    p.add_argument("--ref-metrics", action="store_true",
                   help="режим A: дополнительно считать reference-метрики эталон↔кандидат "
                        "(ROUGE-L всегда; BERTScore при установленном bert-score — "
                        "pip install -r requirements-metrics.txt)")
    p.add_argument("--self-check", action="store_true", help="прогнать self-check и выйти")
    return p


# ══════════════════════════════════════════════════════════════════
# САМОПРОВЕРКА
# ══════════════════════════════════════════════════════════════════

def _force_utf8_stdout() -> None:
    """Windows-консоль по умолчанию cp1251 и падает на кириллице/символах —
    переключаем поток на UTF-8 (как в остальных диагностических прогонах)."""
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass


def _self_check() -> None:
    _force_utf8_stdout()
    logging.basicConfig(level=logging.INFO, format="%(message)s")

    # 1. Легенда непротиворечива по структуре.
    assert CRITICAL_TYPES & BENIGN_TYPES == frozenset(), "тип не может быть одновременно critical и benign"
    print(f"--- легенда: {len(ROW_DISTORTIONS)} строк, "
          f"{len(CRITICAL_TYPES)} critical-типов, {len(BENIGN_TYPES)} benign-типов ---")

    # 2. Грубый классификатор — на эталонных примерах паттернов набора.
    samples = [
        ("число", "Пациент 56 лет.", "Пациент 46 лет."),
        ("отрицание", "Тест положительный.", "Тест не положительный."),
        ("опечатка", "Пациентка госпитализирована.", "Пациентка госпиталлизирована."),
        ("перестановка", "слабость, одышку и боль", "одышку, слабость и боль"),
    ]
    print("--- грубый классификатор (кросс-чек) ---")
    for expected, a, b in samples:
        got = classify_distortion_coarse(a, b)
        mark = "OK " if expected in got else "ERR"
        print(f"  [{mark}] ожидали~{expected:<14} получили={got}")
        assert expected in got, f"классификатор: ожидали {expected} в {got}"

    # 3. Загрузка + guard на реальном наборе (если найден).
    try:
        cases = load_dataset()
    except FileNotFoundError as exc:
        print(f"\n(пропуск прогона на наборе: {exc})\nВСЕ ПРОВЕРКИ КЛАССИФИКАТОРА ПРОЙДЕНЫ")
        return
    assert len(cases) == 10, f"ожидалось 10 случаев, получено {len(cases)}"
    for case in cases:
        assert case.source and case.gold, f"{case.column}: пустой исходник/эталон"
        assert len(case.candidates) >= 30, f"{case.column}: мало кандидатов ({len(case.candidates)})"
    warnings = check_legend_consistency(cases)
    print(f"\n--- guard консистентности: {len(warnings)} предупреждений ---")
    for w in warnings:
        print(f"  ! {w}")

    pairs_a = build_pairs(cases, reference="gold")
    print(f"\n--- режим A: {len(pairs_a)} пар (эталон->кандидат) ---")
    assert len(pairs_a) == sum(len(c.candidates) for c in cases), (
        "режим A не должен содержать контрольных пар «эталон vs эталон»: они по "
        "построению никогда не флагаются и бесплатно раздувают benign-знаменатель")

    # ── Контрольная пара режима B: эталон против самого себя ──────
    pairs_b = build_pairs(cases, reference="gold", with_gold_control=True)
    gold_pairs = [p for p in pairs_b if p.row == GOLD_ROW]
    assert len(gold_pairs) == len(cases), "контрольная пара должна быть у каждого пациента"
    assert all(p.reference == p.candidate for p in gold_pairs), "контроль обязан сравнивать эталон с собой"
    assert gold_pairs[0].spec.category_expected == CATEGORY_READY
    assert gold_pairs[0].spec.e1_expected is False
    print(f"--- режим B: {len(pairs_b)} пар, из них {len(gold_pairs)} контрольных "
          f"(эталон vs эталон; ожидается «{CATEGORY_READY}» без E1) ---")

    # ── Полнота ожиданий ground truth ─────────────────────────────
    for row, spec in sorted(ROW_DISTORTIONS.items()):
        assert spec.category_expected in (CATEGORY_READY, CATEGORY_EDIT, CATEGORY_REJECT), \
            f"строка {row}: неизвестная ожидаемая категория {spec.category_expected!r}"
        assert isinstance(spec.e1_expected, bool), f"строка {row}: ожидание E1 не задано"
    n_e1 = sum(1 for sp in ROW_DISTORTIONS.values() if sp.e1_expected)
    print(f"--- ground truth: {len(ROW_DISTORTIONS)} строк, "
          f"E1 ожидается в {n_e1}, не ожидается в {len(ROW_DISTORTIONS) - n_e1} ---")

    # 4. Быстрый rule-based прогон детекции (без LLM) — дымовой тест.
    det = summarize_detection(run_detection(pairs_a, panel=None))
    print(render_detection(det, llm_used=False))
    print("\nВСЕ ПРОВЕРКИ ПРОЙДЕНЫ (легенда + классификатор + загрузка + rule-based детекция)")


if __name__ == "__main__":
    args = _build_argparser().parse_args()
    if args.self_check:
        _self_check()
        sys.exit(0)
    _force_utf8_stdout()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(name)s: %(message)s")
    import gate
    scope = args.scope or None
    gate.validate_scope(scope)
    run(mode=args.mode, limit_cols=args.limit_cols, use_llm=not args.no_llm,
        profile=args.profile, xlsx_path=args.xlsx, excel_out=args.excel_out,
        ref_metrics=args.ref_metrics, scope=scope)
