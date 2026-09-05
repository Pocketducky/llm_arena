"""
report.py — Блок 5 (продолжение): формирование Excel-отчёта под новую модель
данных — раздельные оси таксономии A-E + детерминированная категория из
aggregator.py, вместо единого `final_score`/`criteria`/`quality`
(старая структура — `save_results` прежней монолитной версии).

Сохранена удобная структура листов прежнего отчёта (сводная/детали/
статистика/критические ошибки), но содержание каждого подчинено новой
модели:
  • «Сводная»     — итоговая категория (код, не самооценка LLM), решение
                    шлюза, статус каждого блока A-E, флаг E1 и его источники,
                    категория по версии LLM-агрегатора — для прямого сравнения
  • «Детали по блокам» — по каждой паре/судье/блоку: какие именно подкритерии
                    не прошли мажоритарную проверку (а не агрегированный балл
                    «на глаз», как раньше)
  • «Статистика»  — по моделям: распределение по 3 категориям, частота
                    срабатывания E1, решения шлюза, согласие код↔LLM
                    (метрика для Блока 6/7 — насколько нужен детерминированный
                    слой поверх самооценки модели)
  • «Критические ошибки» — все срабатывания E1 (с источниками и конкретными
                    примерами из danger_examples) + все блоки со статусом
                    "issues" (с именно теми подкритериями, что не прошли)

Принимает результаты, уже прошедшие aggregator.finalize() — то есть с
полями category/verdict/blocks/e1_triggered/e1_sources/llm_category/decision_path.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import aggregator
import config
import judge

CATEGORY_COLORS = {
    aggregator.CATEGORY_READY:      "C6EFCE",
    aggregator.CATEGORY_EDIT:       "FFEB9C",
    aggregator.CATEGORY_REJECT:     "FFC7CE",
    # Служебные статусы намеренно НЕЙТРАЛЬНОГО цвета: это отсутствие результата,
    # а не плохой результат, и путать их с клиническим отказом нельзя.
    aggregator.CATEGORY_INCOMPLETE: "BDD7EE",
    aggregator.CATEGORY_ERROR:      "D9D9D9",
}

BLOCK_STATUS_COLORS = {"ok": "C6EFCE", "issues": "FFC7CE",
                       "no_data": "BDD7EE", None: "D9D9D9"}

BLOCK_STATUS_RU = {"ok": "ок", "issues": "провал", "no_data": "нет данных"}

_BLOCKS = list(judge.TAXONOMY)   # ("A","B","C","D","E") — порядок таксономии


# ══════════════════════════════════════════════════════════════════
# СТИЛИ (минимальный набор — самодостаточный, без зависимости от evaluator.py,
# который план полностью выводит из эксплуатации)
# ══════════════════════════════════════════════════════════════════

def _brd():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _h(ws, r, c, v, bg="1F4E79", fc="FFFFFF"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=fc, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _brd()


def _d(ws, r, c, v, bold=False, bg=None, red=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, color="C00000" if red else "000000", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _brd()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)


# ══════════════════════════════════════════════════════════════════
# Лист 1 — Сводная таблица
# ══════════════════════════════════════════════════════════════════

def _sheet_summary(wb: Workbook, results: list[dict]):
    ws = wb.active
    ws.title = "Сводная таблица"
    hdrs = (["ЭМК", "Модель", "scope", "Категория\n(итог)", "🚨 E1", "Источники E1",
             "E1: цитата\nподтверждена?", "Шлюз (Блок 3)", "Причины шлюза",
             "Объективный слой\n(Блок 2)"]
            + [f"Блок {b}" for b in _BLOCKS]
            + ["Блоков\nбез данных", "Обрывов\nответа", "Спасено\nсалважем", "Таймаутов",
               "Время,\nс", "Категория\nпо LLM", "Совпадает\nс LLM?", "Вердикт"])
    for ci, h in enumerate(hdrs, 1):
        _h(ws, 1, ci, h)
    ws.row_dimensions[1].height = 48
    ws.freeze_panes = "A2"

    for ridx, r in enumerate(results, 2):
        category = r.get("category", "ошибка")
        blocks = r.get("blocks", {}) or {}
        gate = r.get("gate", {}) or {}
        objective = r.get("objective")
        e1 = bool(r.get("e1_triggered"))
        sources = ", ".join(r.get("e1_sources", []) or []) or "—"
        llm_cat = r.get("llm_category") or "—"
        agree = (category == r.get("llm_category")) if r.get("llm_category") else None

        tele = r.get("telemetry") or {}
        e1sig = r.get("e1_signals") or {}
        nodata = r.get("nodata_blocks") or [b for b, v in blocks.items()
                                            if (v or {}).get("status") == "no_data"]
        gate_codes = ", ".join(gate.get("reasons") or []) or "—"

        row = ([r["emr_id"], r["model_id"], r.get("scope") or "—", category,
                "🚨 ДА" if e1 else "нет", sources, _e1_citation_cell(e1sig),
                gate.get("status", "—"), gate_codes, _objective_cell_text(objective)]
               + [_block_cell_text(blocks.get(b)) for b in _BLOCKS]
               + [len(nodata) or "—",
                  tele.get("truncated", 0) + tele.get("truncation_repaired", 0) or "—",
                  tele.get("salvaged_keys", 0) or "—",
                  tele.get("timeouts", 0) or "—",
                  tele.get("wall_seconds", "—"),
                  llm_cat,
                  ("да" if agree else ("нет" if agree is False else "—")),
                  r.get("verdict", "")[:200]])

        bgs = ([None, None, None, CATEGORY_COLORS.get(category, "FFFFFF"),
                "FFC7CE" if e1 else None, None,
                ("FFC7CE" if e1sig.get("disputed_by_aggregator") else None),
                _gate_color(gate.get("status")), None,
                ("FFC7CE" if _objective_has_findings(objective) else None)]
               + [BLOCK_STATUS_COLORS.get((blocks.get(b) or {}).get("status")) for b in _BLOCKS]
               + [("BDD7EE" if nodata else None),
                  ("FFEB9C" if tele.get("truncated") or tele.get("truncation_repaired") else None),
                  ("FFEB9C" if tele.get("salvaged_keys") else None),
                  ("FFEB9C" if tele.get("timeouts") else None),
                  None, None, ("FFC7CE" if agree is False else None), None])

        for ci, (val, bg) in enumerate(zip(row, bgs), 1):
            _d(ws, ridx, ci, val, bold=(ci == 4), bg=bg)
        ws.row_dimensions[ridx].height = 30

    widths = ([10, 9, 12, 16, 7, 16, 14, 12, 24, 26] + [9] * len(_BLOCKS)
              + [9, 9, 9, 9, 8, 16, 9, 50])
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _block_cell_text(block: Optional[dict]) -> str:
    """Ячейка блока A-E.

    Важно отличать «провалено» от «нет данных»: раньше оба случая рисовались
    одним значком «✗», и строка «Балл 100 %, статус issues, не прошли —»
    (частично разобранный ответ судьи) выглядела как претензия без претензии.
    """
    if not block:
        return "—"
    status = block.get("status")
    score = block.get("score")
    failed = block.get("failed_subcriteria") or []
    undet = block.get("undetermined_subcriteria") or []

    if status == "no_data":
        return f"❓ н/д ({', '.join(undet)})" if undet else "❓ н/д"
    mark = "✓" if status == "ok" else "✗"
    label = f"{mark} {score:.0f}%" if score is not None else f"{mark} —"
    if failed:
        label += f" ({', '.join(failed)})"
    if undet:
        label += f" [н/д: {', '.join(undet)}]"
    return label


def _gate_color(status) -> Optional[str]:
    return {"pass": "C6EFCE", "rework": "FFEB9C", "reject": "FFC7CE"}.get(status)


def _e1_citation_cell(e1sig: dict) -> str:
    """Подтверждён ли E1 проверяемой цитатой из суммаризации.

    В v1 стоп-правило было дизъюнкцией без доказательства и срабатывало в 38
    случаях из 38, включая эталон. Теперь читателю отчёта видно, чем именно
    подкреплён каждый флаг — и какие флаги отброшены как неподтверждённые.
    """
    if not e1sig:
        return "—"
    if not e1sig.get("require_citation", True):
        return "проверка отключена"
    if e1sig.get("disputed_by_aggregator"):
        return "✗ только агрегатор, судьи не подтвердили"
    confirmed = e1sig.get("raised_by_judges") or []
    without = e1sig.get("raised_without_citation") or []
    if confirmed:
        quotes = e1sig.get("citations") or {}
        n = sum(len(v) for v in quotes.values())
        return f"✓ да ({n} цитат)"
    if without:
        return f"✗ флаг без цитаты: {', '.join(without)}"
    return "—"


# ── Объективный слой (Блок 2/7) — компактное отображение «рядом с LLM» ──
# Принимает уже сериализованное резюме (objective_layer.
# ObjectiveComparisonReport.to_summary(), прокинутое через judge.
# evaluate_summary -> aggregator.finalize в поле "objective"). None
# означает «не построено» — честно (пара отклонена шлюзом ДО этого шага,
# см. комментарий в judge.evaluate_summary про экономию GPU-времени),
# а не «забыли посчитать» — поэтому здесь не ошибка, а отдельное «—».

def _objective_cell_text(objective: Optional[dict]) -> str:
    if not objective:
        return "— (шлюз отклонил до сверки)"
    num = objective.get("numeric") or {}
    pol = objective.get("polarity") or {}
    parts = []

    mm, um, total_num = num.get("mismatch_count", 0), num.get("unit_mismatch_count", 0), num.get("total_in_a", 0)
    if mm or um:
        bits = []
        if mm:
            bits.append(f"{mm} знач.")
        if um:
            bits.append(f"{um} ед.изм.")
        parts.append(f"числа: {' + '.join(bits)} расхожд. (из {total_num})")

    fl, total_pol = pol.get("flip_count", 0), pol.get("total_in_a", 0)
    if fl:
        parts.append(f"полярность: {fl} инверс. (из {total_pol})")

    ents = objective.get("entities")
    if ents:
        f1s = [rep["f1"] for rep in ents.values()]
        if f1s:
            parts.append(f"сущности: F1≈{sum(f1s) / len(f1s):.2f}")

    return "; ".join(parts) if parts else "✓ расхождений не найдено"


def _objective_has_findings(objective: Optional[dict]) -> bool:
    """«Жёсткие» автоматические находки — то, что обязано бросаться в глаза
    рядом с оценкой судей (числовые/полярные расхождения, введённая ложная
    причинность — см. Блок 2)."""
    if not objective:
        return False
    num = objective.get("numeric") or {}
    pol = objective.get("polarity") or {}
    cau = objective.get("causality") or {}
    return bool(num.get("mismatch_count") or num.get("unit_mismatch_count")
                or pol.get("flip_count") or cau.get("introduced"))


# ══════════════════════════════════════════════════════════════════
# Лист 2 — Детали по блокам и судьям (что именно не прошло и почему)
# ══════════════════════════════════════════════════════════════════

def _sheet_block_details(wb: Workbook, results: list[dict]):
    ws = wb.create_sheet("Детали по блокам")
    hdrs = ["ЭМК", "Модель", "Блок", "Балл, %", "Статус",
            "Не прошли (мажоритарно)", "Без данных", "Сбой разбора у судей",
            "Голоса судей по подкритериям", "Пересмотр R1→R2",
            "Комментарии судей (по непрошедшим)"]
    for ci, h in enumerate(hdrs, 1):
        _h(ws, 1, ci, h)
    ws.freeze_panes = "A2"

    ridx = 2
    for r in results:
        blocks = r.get("blocks", {}) or {}
        reports = aggregator._final_reports(r.get("r1", {}), r.get("r2", {}))
        for b in _BLOCKS:
            v = blocks.get(b)
            if v is None:
                continue
            failed = v.get("failed_subcriteria") or []
            comments = _collect_comments(reports, b, failed)
            row = [r["emr_id"], r["model_id"], b,
                   f"{v['score']:.0f}" if v.get("score") is not None else "—",
                   BLOCK_STATUS_RU.get(v["status"], v["status"]),
                   ", ".join(failed) or "—",
                   ", ".join(v.get("undetermined_subcriteria") or []) or "—",
                   ", ".join(v.get("parse_error_roles") or []) or "—",
                   _votes_cell(reports, b),
                   _revision_cell(r.get("r1", {}), r.get("r2", {}), b),
                   comments[:300]]
            bg = BLOCK_STATUS_COLORS.get(v["status"])
            for ci, val in enumerate(row, 1):
                _d(ws, ridx, ci, val, bg=(bg if ci == 5 else None))
            ridx += 1

    for i, w in enumerate([10, 9, 7, 8, 10, 22, 16, 18, 40, 26, 60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _votes_cell(reports: dict[str, dict], block: str) -> str:
    """Голос КАЖДОГО судьи по каждому подкритерию блока.

    Раньше наружу выходил только мажоритарный итог, поэтому «2 из 3 против 1» и
    «3 из 3» выглядели одинаково, а расхождение судей — главный сигнал спорности
    случая — из отчёта не читалось вовсе."""
    codes = judge.TAXONOMY[block]
    out = []
    for code in codes:
        marks = []
        for role in sorted(reports):
            sub = (reports[role].get(block) or {}).get(code)
            if isinstance(sub, dict) and isinstance(sub.get("pass"), bool):
                marks.append("✓" if sub["pass"] else "✗")
            else:
                marks.append("?")
        out.append(f"{code}:{''.join(marks)}")
    return "  ".join(out)


def _revision_cell(r1: dict, r2: dict, block: str) -> str:
    """Что судьи изменили в себе после cross-peer-review (R1 -> R2).

    Смысл раунда 2 — содержательный пересмотр. Если он не меняет НИЧЕГО, раунд
    стоит своих LLM-вызовов впустую, и это должно быть видно в отчёте, а не
    предполагаться."""
    changes = []
    for role, before in (r1 or {}).items():
        after = (r2 or {}).get(role)
        if not isinstance(before, dict) or not isinstance(after, dict):
            continue
        for code in judge.TAXONOMY[block]:
            b = (before.get(block) or {}).get(code)
            a = (after.get(block) or {}).get(code)
            if not isinstance(b, dict) or not isinstance(a, dict):
                continue
            if b.get("pass") != a.get("pass"):
                changes.append(f"{role}/{code}: {b.get('pass')}→{a.get('pass')}")
    return "; ".join(changes) or "без изменений"


def _collect_comments(reports: dict[str, dict], block: str, failed_codes: list[str]) -> str:
    """Конкретные обоснования судей по подкритериям, не прошедшим мажоритарную
    проверку — то самое содержательное «почему», которое старый compact()
    обрезал перед обменом между судьями (см. judge._format_full_report);
    здесь оно доходит до читателя отчёта в явном виде."""
    parts = []
    for code in failed_codes:
        for role, rep in reports.items():
            sub = rep.get(block, {}).get(code)
            if isinstance(sub, dict) and sub.get("pass") is False and sub.get("comment"):
                parts.append(f"{code}/{role}: {sub['comment']}")
    return "; ".join(parts)


# ══════════════════════════════════════════════════════════════════
# Лист 3 — Статистика по моделям
# ══════════════════════════════════════════════════════════════════

def _sheet_statistics(wb: Workbook, results: list[dict]):
    """Статистика по моделям.

    ГЛАВНОЕ ИСПРАВЛЕНИЕ: колонки категорий теперь СХОДЯТСЯ С N. В v1 выводились
    только три клинические категории, а строки со статусом «ошибка» увеличивали
    N и не попадали никуда — суммы молча не сходились, и доля отказов читалась
    завышенной. Служебные статусы («Оценка неполна», «ошибка») показаны
    отдельными колонками: это отсутствие результата, а не плохой результат.
    """
    ws = wb.create_sheet("Статистика")
    hdrs = ["Модель", "N",
            aggregator.CATEGORY_READY, aggregator.CATEGORY_EDIT, aggregator.CATEGORY_REJECT,
            aggregator.CATEGORY_INCOMPLETE, "ошибка", "Сумма = N?",
            "🚨 E1, %", "E1 с цитатой, %", "E1 оспорен\nагрегатором",
            "Шлюз: pass", "Шлюз: rework", "Шлюз: reject",
            "Согласие код↔LLM, %",
            "Объективный слой (Блок 2) — доля пар с расхождениями:",
            "  числа/ед.изм., %", "  полярность, %"]
    for ci, h in enumerate(hdrs, 1):
        _h(ws, 1, ci, h)
    ws.row_dimensions[1].height = 46
    # Заголовок-«зонтик» над двумя последними колонками.
    ws.merge_cells(start_row=1, start_column=16, end_row=1, end_column=18)

    df = pd.DataFrame([{
        "model_id": r["model_id"],
        "category": r.get("category", aggregator.CATEGORY_ERROR),
        "e1": bool(r.get("e1_triggered")),
        "e1_cited": bool(((r.get("e1_signals") or {}).get("citations")) or {}),
        "e1_disputed": bool((r.get("e1_signals") or {}).get("disputed_by_aggregator")),
        "gate": (r.get("gate") or {}).get("status"),
        "agree": (r.get("category") == r.get("llm_category")) if r.get("llm_category") else None,
        # Доля пар, где объективный слой нашёл хоть одно числовое/полярное
        # расхождение — None для записей без objective: для них вопрос
        # «было ли расхождение» неприменим, а не «нет».
        "obj_numeric_finding": _obj_flag(r.get("objective"), "numeric"),
        "obj_polarity_finding": _obj_flag(r.get("objective"), "polarity"),
    } for r in results])

    if not df.empty:
        for ridx, (mid, g) in enumerate(df.groupby("model_id"), 2):
            cc = g["category"].value_counts()
            agree = g["agree"].dropna()
            num_flags = g["obj_numeric_finding"].dropna()
            pol_flags = g["obj_polarity_finding"].dropna()
            counts = [int(cc.get(c, 0)) for c in
                      (aggregator.CATEGORY_READY, aggregator.CATEGORY_EDIT,
                       aggregator.CATEGORY_REJECT, aggregator.CATEGORY_INCOMPLETE,
                       aggregator.CATEGORY_ERROR)]
            balanced = sum(counts) == len(g)
            e1_rows = g[g["e1"]]
            row = ([mid, len(g)] + counts
                   + ["да" if balanced else f"НЕТ ({sum(counts)} против {len(g)})",
                      round(100 * g["e1"].mean(), 1),
                      round(100 * e1_rows["e1_cited"].mean(), 1) if len(e1_rows) else None,
                      int(g["e1_disputed"].sum()),
                      int((g["gate"] == "pass").sum()),
                      int((g["gate"] == "rework").sum()),
                      int((g["gate"] == "reject").sum()),
                      round(100 * agree.mean(), 1) if len(agree) else None,
                      None,   # «зонтик»
                      round(100 * num_flags.mean(), 1) if len(num_flags) else None,
                      round(100 * pol_flags.mean(), 1) if len(pol_flags) else None])
            for ci, v in enumerate(row, 1):
                bg = None
                if ci == 5 and isinstance(v, int) and v > 0:            # «Неприемлемо»
                    bg = "FFC7CE"
                elif ci in (6, 7) and isinstance(v, int) and v > 0:     # служебные статусы
                    bg = "BDD7EE"
                elif ci == 8 and not balanced:
                    bg = "FFC7CE"
                elif ci in (17, 18) and isinstance(v, (int, float)) and v > 0:
                    bg = "FFC7CE"
                _d(ws, ridx, ci, v, bg=bg)

    for i, w in enumerate([14, 6, 16, 16, 14, 14, 9, 14, 9, 12, 12, 11, 12, 12, 14, 4, 14, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _obj_flag(objective: Optional[dict], section: str) -> Optional[bool]:
    """True/False — «нашёл ли объективный слой расхождение в данном
    разрезе (numeric/polarity)»; None — резюме недоступно (пара отклонена
    шлюзом до сверки, см. _objective_cell_text). None исключается из
    усреднения в _sheet_statistics — доля считается по применимым парам."""
    if not objective:
        return None
    blk = objective.get(section) or {}
    if section == "numeric":
        return bool(blk.get("mismatch_count") or blk.get("unit_mismatch_count"))
    if section == "polarity":
        return bool(blk.get("flip_count"))
    return None


# ══════════════════════════════════════════════════════════════════
# Лист 4 — Критические ошибки (E1 + проблемные блоки)
# ══════════════════════════════════════════════════════════════════

def _sheet_critical(wb: Workbook, results: list[dict]):
    ws = wb.create_sheet("🚨 Критические")
    hdrs = ["ЭМК", "Модель", "Тип", "Блок/подкритерий", "Описание",
            "Цитата из суммаризации", "Подтверждена кодом?", "Категория"]
    for ci, h in enumerate(hdrs, 1):
        _h(ws, 1, ci, h, bg="C00000")

    ridx = 2
    for r in results:
        category = r.get("category", "ошибка")
        e1sig = r.get("e1_signals") or {}
        if r.get("e1_triggered"):
            sources = ", ".join(r.get("e1_sources", []) or []) or "—"
            examples = _collect_e1_examples(aggregator._final_reports(r.get("r1", {}), r.get("r2", {})))
            quotes = "; ".join(q for v in (e1sig.get("citations") or {}).values() for q in v)
            row = [r["emr_id"], r["model_id"], "🚨 E1 — стоп-правило", f"E1 (источники: {sources})",
                   (examples or r.get("verdict", ""))[:250],
                   quotes[:250] or "—",
                   "✓ да" if quotes else "нет цитаты (совпадение сигналов)", category]
            for ci, v in enumerate(row, 1):
                _d(ws, ridx, ci, v, bg="FFC7CE", bold=True)
            ridx += 1

        # Отброшенные флаги E1 — не менее важны для доверия к предохранителю:
        # видно, что система не просто «не сработала», а осознанно не засчитала
        # флаг без доказательства.
        for role, bad in (e1sig.get("unverifiable_citations") or {}).items():
            row = [r["emr_id"], r["model_id"], "E1 отклонён (цитата не найдена)",
                   f"E1 / {role}", "; ".join(str(x) for x in bad)[:250],
                   "; ".join(str(x) for x in bad)[:250],
                   "✗ фрагмента нет в суммаризации", category]
            for ci, v in enumerate(row, 1):
                _d(ws, ridx, ci, v, bg="DDEBF7")
            ridx += 1
        if e1sig.get("disputed_by_aggregator"):
            row = [r["emr_id"], r["model_id"], "E1 оспорен",
                   "E1 / агрегатор R3",
                   "Агрегатор поднял стоп-правило, но ни один судья его не поднял",
                   "—", "✗ нет подтверждения судьями", category]
            for ci, v in enumerate(row, 1):
                _d(ws, ridx, ci, v, bg="DDEBF7")
            ridx += 1

        blocks = r.get("blocks", {}) or {}
        reports = aggregator._final_reports(r.get("r1", {}), r.get("r2", {}))
        for b, v in blocks.items():
            if v.get("status") != "issues":
                continue
            for code in (v.get("failed_subcriteria") or []):
                comment = next((reports[role][b][code]["comment"]
                                for role in reports
                                if reports[role].get(b, {}).get(code, {}).get("pass") is False
                                and reports[role][b][code].get("comment")), "")
                row = [r["emr_id"], r["model_id"], "Замечание по таксономии",
                       f"{b}/{code} — {judge.SUBCRITERIA_LABELS_RU.get(code, '')}",
                       comment[:250], "—", "—", category]
                for ci, val in enumerate(row, 1):
                    _d(ws, ridx, ci, val, bg="FFEB9C")
                ridx += 1

        # Блоки без данных раньше не попадали сюда вовсе: лист перебирал только
        # failed_subcriteria, поэтому подкритерии, РЕАЛЬНО вызвавшие отказ через
        # «нет данных», не показывались нигде.
        for b, v in blocks.items():
            if v.get("status") != "no_data":
                continue
            row = [r["emr_id"], r["model_id"], "Нет данных (сбой разбора)",
                   f"{b}: {', '.join(v.get('undetermined_subcriteria') or [])}",
                   ("Судьи не вернули разбираемых оценок"
                    + (f" (сбой у: {', '.join(v.get('parse_error_roles') or [])})"
                       if v.get("parse_error_roles") else "")),
                   "—", "—", category]
            for ci, val in enumerate(row, 1):
                _d(ws, ridx, ci, val, bg="BDD7EE")
            ridx += 1

    for i, w in enumerate([10, 12, 26, 36, 60, 40, 26, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _collect_e1_examples(reports: dict[str, dict]) -> str:
    parts = []
    for role, rep in reports.items():
        e1 = rep.get("E", {}).get("E1")
        if isinstance(e1, dict) and e1.get("pass") is False:
            ex = "; ".join(str(x) for x in (e1.get("danger_examples") or []))
            comment = e1.get("comment", "")
            parts.append(f"{role}: {comment}" + (f" [{ex}]" if ex else ""))
    return " | ".join(parts)


# ══════════════════════════════════════════════════════════════════
# Лист 5 — Трасса решений: ответ на вопрос «почему такой вердикт»
# ══════════════════════════════════════════════════════════════════
# aggregator.finalize строит пошаговое, человекочитаемое обоснование
# (decision_path) — и в v1 его НЕ ЧИТАЛ никто, включая этот модуль. Это ровно
# то, чего не хватало читателю отчёта: видно категорию, но не видно, каким
# правилом она получена.

def _sheet_decision_path(wb: Workbook, results: list[dict]):
    ws = wb.create_sheet("Трасса решений")
    hdrs = ["ЭМК", "Модель", "Категория", "Шаг", "Обоснование"]
    for ci, h in enumerate(hdrs, 1):
        _h(ws, 1, ci, h)
    ws.freeze_panes = "A2"

    ridx = 2
    for r in results:
        category = r.get("category", aggregator.CATEGORY_ERROR)
        path = r.get("decision_path") or ["(трасса не сохранена)"]
        for step_no, step in enumerate(path, 1):
            if not str(step).strip():
                continue
            row = [r["emr_id"], r["model_id"], category, step_no, str(step)]
            bg = CATEGORY_COLORS.get(category) if step_no == 1 else None
            for ci, val in enumerate(row, 1):
                _d(ws, ridx, ci, val, bg=(bg if ci == 3 else None))
            ws.cell(row=ridx, column=5).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)
            ridx += 1

    for i, w in enumerate([10, 9, 20, 6, 120], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════
# Лист 6 — Диагностика прогона: надёжность отдельно от качества
# ══════════════════════════════════════════════════════════════════
# Обрывы ответов, ретраи, таймауты, салважи, токены и латентность в v1
# существовали ТОЛЬКО как строки в stdout. По итоговому Excel нельзя было
# отличить «судьи нашли проблемы» от «мы не смогли разобрать ответ судьи» —
# при том, что второе напрямую превращалось в клиническую категорию.

_TELE_COLUMNS = [
    ("calls", "Вызовов\nк модели"),
    ("attempts", "Попыток\nask_json"),
    ("truncated", "Обрывов\nпо лимиту"),
    ("truncation_repaired", "Достроено\nпосле обрыва"),
    ("salvaged", "Салважей"),
    ("salvaged_keys", "Полей\nспасено"),
    ("budget_raised", "Подъёмов\nбюджета"),
    ("json_failures", "Провалов\nразбора"),
    ("fallback_used", "Ответов без\nзаземления"),
    ("timeouts", "Таймаутов"),
    ("http_errors", "HTTP-ошибок"),
    ("conn_errors", "Обрывов\nсвязи"),
    ("prompt_tokens", "Токенов\nпромпта"),
    ("completion_tokens", "Токенов\nответа"),
    ("wall_seconds", "Время,\nс"),
]


def _sheet_diagnostics(wb: Workbook, results: list[dict]):
    ws = wb.create_sheet("Диагностика прогона")
    hdrs = ["ЭМК", "Модель", "Категория"] + [t for _, t in _TELE_COLUMNS] + ["События"]
    for ci, h in enumerate(hdrs, 1):
        _h(ws, 1, ci, h)
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "C2"

    totals = {k: 0 for k, _ in _TELE_COLUMNS}
    ridx = 2
    for r in results:
        tele = r.get("telemetry") or {}
        vals = []
        for key, _ in _TELE_COLUMNS:
            v = tele.get(key, 0) or 0
            totals[key] += v if isinstance(v, (int, float)) else 0
            vals.append(v)
        notes = "; ".join(str(n) for n in (tele.get("notes") or []))[:400]
        row = [r["emr_id"], r["model_id"], r.get("category", "—")] + vals + [notes or "—"]
        for ci, val in enumerate(row, 1):
            key = _TELE_COLUMNS[ci - 4][0] if 4 <= ci <= 3 + len(_TELE_COLUMNS) else None
            bad = key in ("truncated", "truncation_repaired", "salvaged", "timeouts",
                          "http_errors", "conn_errors", "fallback_used", "json_failures")
            bg = "FFEB9C" if (bad and isinstance(val, (int, float)) and val) else None
            _d(ws, ridx, ci, val, bg=bg)
        ridx += 1

    if results:
        _d(ws, ridx, 1, "ИТОГО", bold=True)
        _d(ws, ridx, 2, "", bold=True)
        _d(ws, ridx, 3, f"{len(results)} пар", bold=True)
        for ci, (key, _) in enumerate(_TELE_COLUMNS, 4):
            v = round(totals[key], 1) if key == "wall_seconds" else totals[key]
            _d(ws, ridx, ci, v, bold=True,
               bg=("FFC7CE" if key in ("truncated", "timeouts", "conn_errors") and v else None))

    for i, w in enumerate([10, 9, 18] + [11] * len(_TELE_COLUMNS) + [70], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════
# Лист 7 — Пропущенные пары: что НЕ оценено и почему
# ══════════════════════════════════════════════════════════════════

def _sheet_skipped(wb: Workbook, results: list[dict]):
    """Пары со служебным статусом. Их нельзя молча растворять в статистике:
    это перечень работы, которую нужно доделать перепрогоном."""
    skipped = [r for r in results
               if r.get("category") in aggregator.SERVICE_CATEGORIES]
    ws = wb.create_sheet("Пропущенные пары")
    hdrs = ["ЭМК", "Модель", "Статус", "Что произошло", "Блоки без данных",
            "Стадия/события", "Что делать"]
    for ci, h in enumerate(hdrs, 1):
        _h(ws, 1, ci, h, bg="7F6000")
    ws.freeze_panes = "A2"

    if not skipped:
        _d(ws, 2, 1, "Пропущенных пар нет — все пары оценены полностью.", bg="C6EFCE")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
        for i, w in enumerate([10, 9, 18, 60, 20, 60, 40], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return

    for ridx, r in enumerate(skipped, 2):
        tele = r.get("telemetry") or {}
        status = r.get("category")
        nodata = ", ".join(r.get("nodata_blocks") or []) or "—"
        if status == aggregator.CATEGORY_ERROR:
            action = ("Перепрогнать пару: сбой инфраструктуры, а не качество "
                      "суммаризации. Чекпоинт не сохранён — повторный запуск "
                      "возьмёт её заново.")
        else:
            action = ("Перепрогнать пару: ответы судей не разобраны. Если "
                      "повторяется — поднять бюджет токенов соответствующего "
                      "раунда (EMR_MAX_TOKENS_R1/R2/R3).")
        row = [r["emr_id"], r["model_id"], status,
               (r.get("verdict") or "")[:400], nodata,
               "; ".join(str(n) for n in (tele.get("notes") or []))[:400] or "—",
               action]
        for ci, val in enumerate(row, 1):
            _d(ws, ridx, ci, val, bg=(CATEGORY_COLORS.get(status) if ci == 3 else None))

    for i, w in enumerate([10, 9, 18, 60, 20, 60, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════
# Лист 8 — Конфигурация прогона: чем именно получены эти цифры
# ══════════════════════════════════════════════════════════════════

def _sheet_config(wb: Workbook, results: list[dict], run_meta: Optional[dict] = None):
    """Полный срез параметров прогона.

    Без него результаты нельзя ни воспроизвести, ни сравнить между прогонами:
    та же таблица при другом пороге, бюджете токенов или seed означает другое.
    Данные берутся из audit.versions_snapshot()/_thresholds_snapshot() — они
    уже собирались для JSONL, но в Excel не попадали никогда.
    """
    import audit as _audit

    ws = wb.create_sheet("Конфигурация прогона")
    for ci, h in enumerate(["Параметр", "Значение"], 1):
        _h(ws, 1, ci, h)

    meta = dict(run_meta or {})
    rows: list[tuple] = [("── ПРОГОН ──", "")]
    for key in ("run_id", "profile", "scope", "concurrency", "n_pairs",
                "dataset", "started", "elapsed"):
        if key in meta:
            rows.append((key, meta[key]))
    rows.append(("пар в отчёте", len(results)))

    try:
        versions = _audit.versions_snapshot(meta.get("profile"))
    except Exception as exc:   # noqa: BLE001 — отчёт не должен падать из-за среза
        versions = {"ошибка среза версий": str(exc)}

    rows.append(("── МОДЕЛИ И РОЛИ ──", ""))
    for role, model in (versions.get("roles") or {}).items():
        rows.append((role, model))

    rows.append(("── ВЕРСИИ ──", ""))
    for key in ("profile", "prompt_set_version", "taxonomy_version", "decision_table_version"):
        if key in versions:
            rows.append((key, versions[key]))

    rows.append(("── ПОРОГИ И ПАРАМЕТРЫ РЕШЕНИЙ ──", ""))
    for key, val in (versions.get("thresholds") or {}).items():
        rows.append((key, str(val)))

    rows.append(("── БЭКЕНД ──", ""))
    rows.append(("backend", config.LLM_BACKEND))
    if config.LLM_BACKEND == "vllm":
        rows.append(("vllm endpoints", ", ".join(config.vllm_endpoints())))
        rows.append(("max_model_len (обнаружено)", meta.get("max_model_len", "не проверялось")))
    else:
        rows.append(("ollama num_ctx", config.NUM_CTX))
    rows.append(("timeout, с", config.TIMEOUT_SECONDS))

    rows.append(("── КОД ──", ""))
    rows.append(("git commit", _git_commit()))

    for ridx, (k, v) in enumerate(rows, 2):
        header = str(k).startswith("──")
        _d(ws, ridx, 1, k, bold=header, bg=("DDEBF7" if header else None))
        _d(ws, ridx, 2, v, bg=("DDEBF7" if header else None))
        ws.cell(row=ridx, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=ridx, column=2).alignment = Alignment(horizontal="left", vertical="center",
                                                          wrap_text=True)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 80


def _git_commit() -> str:
    """Коммит, которым посчитан отчёт — чтобы цифры можно было привязать к коду."""
    import subprocess
    try:
        out = subprocess.run(["git", "rev-parse", "--short", "HEAD"],
                             cwd=Path(__file__).resolve().parent,
                             capture_output=True, text=True, timeout=5)
        commit = out.stdout.strip()
        dirty = subprocess.run(["git", "status", "--porcelain"],
                               cwd=Path(__file__).resolve().parent,
                               capture_output=True, text=True, timeout=5).stdout.strip()
        return f"{commit}{' (есть незакоммиченные правки)' if dirty else ''}" or "неизвестно"
    except Exception:   # noqa: BLE001 — git может отсутствовать
        return "неизвестно"


# ══════════════════════════════════════════════════════════════════
# ПУБЛИЧНАЯ ТОЧКА ВХОДА
# ══════════════════════════════════════════════════════════════════

def save_results(results: list[dict], path: str, *, run_meta: Optional[dict] = None):
    """results — список словарей, каждый уже прошедший aggregator.finalize().

    run_meta — необязательные метаданные прогона (run_id, профиль, scope,
    длительность, обнаруженный max_model_len) для листа «Конфигурация прогона».

    Восемь листов: четыре прежних (обогащённых) и четыре новых. Новые отвечают
    на вопросы, на которые отчёт v1 не отвечал вовсе:
      • «Трасса решений»       — ПОЧЕМУ выставлена именно эта категория;
      • «Диагностика прогона»  — надёжность прогона отдельно от качества
                                 суммаризаций (обрывы, ретраи, таймауты, токены);
      • «Пропущенные пары»     — что не оценено и что с этим делать;
      • «Конфигурация прогона» — чем именно получены эти цифры.
    """
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _sheet_summary(wb, results)
    _sheet_block_details(wb, results)
    _sheet_statistics(wb, results)
    _sheet_critical(wb, results)
    _sheet_decision_path(wb, results)
    _sheet_diagnostics(wb, results)
    _sheet_skipped(wb, results)
    _sheet_config(wb, results, run_meta)
    wb.save(path)
